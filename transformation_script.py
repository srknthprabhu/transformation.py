import re
import shutil
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import argparse  # For parsing commandLine
import sys
import os
import os.path
import logging
from datetime import datetime
import time
import datetime as dt
from datetime import date
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from Onboarding_script import cinlist
print(cinlist)
def excel_date(date1):
    temp = dt.datetime(1899, 12, 30)  # Note, not 31st Dec but 30th!
    delta = date1 - temp
    return float(delta.days) + (float(delta.seconds) / 86400)
def find_number(text):
    num = re.findall(r"[-+]?\d*\.\d+|\d+", text)
    return float(num[0])
def removeduplicate(array):
    res = []
    for i in array:
        if i not in res:
            res.append(i)
    return res
def start_conversionProcess(fullPath, bFile, fileType):
    print("{}{}{}".format("Start converting Process for <", bFile, ">", fileType))
    location = fullPath  # set location as full path of input files
    wb = load_workbook(location)  # loading workbook
    log = pd.DataFrame()  # creating log sheet
    log.insert(0, 'Sheet Name', np.nan)
    log.insert(1, 'Error At', np.nan)
    log.insert(2, 'Warning At', np.nan)
    log.insert(3, 'Fixed Legal Team', np.nan)
    log.insert(4, 'Warning Hissa', np.nan)
    log.insert(5, 'Fixed Hissa', np.nan)
    logging.basicConfig(format='%(asctime)s %(message)s', datefmt='%Y/%m/%d %I:%M:%S%p')
    pd.set_option('mode.chained_assignment', None)
    User = pd.DataFrame(index=[0])  # Creating User sheet
    User.insert(0, "FullName", np.nan)
    User.insert(1, "Email", np.nan)
    User.insert(2, "Password", np.nan)
    User['FullName'][0] = 'Review Team'
    User['Password'][0] = 'review'  # default password is review
    User['Email'][0] = 'review1@rulezero.com'
    Sheets = pd.read_excel(location, sheet_name=None)  # Reading all sheets
    ky = list(Sheets.keys())  # list of all sheetnames
    for i in range(len(ky)):
        ky[i] = ky[i].lower()
    # Converted into lowercase to use string matching
    for i in range(len(ky)):
        # Finding company for company sheet
        if (ky[i].find("company") != -1):
            Company = pd.read_excel(location, sheet_name=ky.index(ky[i]))
            # skipping the empty row
            if Company.columns.tolist()[1] == 'Unnamed: 1':
                Company = pd.read_excel(location, sheet_name=ky.index(ky[i]), skiprows=1)
            Company.drop(Company.columns[Company.columns.str.contains('Unnamed', case=False)], axis=1, inplace=True)
            company_new = pd.DataFrame()  # new company sheet created
            Company.columns = map(str.lower, Company.columns)  # Lowercase for string matching
            index_value = 0
            input_file = {k: 0 for v, k in enumerate(Company.columns)}
            out_file = []
            for i in range(len(Company.columns)):
                if (Company.columns[i].find("cin") != -1):
                    company_new.insert(index_value, "CIN", Company[Company.columns[i]].tolist())
                    input_file.update({Company.columns[i]: 1})
                    out_file.append("CIN")
                    continue
                elif (Company.columns[i].find("authorised") != -1) or (Company.columns[i].find("authorized") != -1):
                    if (Company.columns[i].find("incorporation")) == -1:
                        company_new.insert(index_value, "Authorised Capital (INR) as of date",      Company[Company.columns[i]].tolist())
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Authorised Capital (INR) as of date")
                        continue
                    if (Company.columns[i].find("incorporation") != -1):
                        company_new.insert(index_value, "Authorised Capital at Incorporation",      Company[Company.columns[i]].tolist())
                        
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Authorised Capital at Incorporation")
                        continue
                elif (Company.columns[i].find("incorporation") != -1):
                    company_new.insert(index_value, "Date of Incorporation", Company[Company.columns[i]].tolist())
                    
                    input_file.update({Company.columns[i]: 1})
                    out_file.append("Date of Incorporation")
                    continue
                elif (Company.columns[i].find("capital") != -1):
                    if (Company.columns[i].find("paid") != -1):
                        company_new.insert(index_value, "Paid up Capital (INR)", Company[Company.columns[i]].tolist())
                        
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Paid up Capital (INR)")
                elif (Company.columns[i].find("company") != -1):
                    if (Company.columns[i].find("name") != -1):
                        company_new.insert(index_value, "Company Name", Company[Company.columns[i]].tolist())

                        
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Company Name")
                        continue
                    elif (Company.columns[i].find("category") != -1):
                        company_new.insert(index_value, "Company Category", Company[Company.columns[i]].tolist())
                        
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Company Category")
                    elif (Company.columns[i].find("class") != -1):
                        company_new.insert(index_value, "Class of Company", Company[Company.columns[i]].tolist())
                        
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Class of Company")
                        continue
                    elif (Company.columns[i].find("efiling") != -1):
                        company_new.insert(index_value, "Company Status(for efiling)",      Company[Company.columns[i]].tolist())
                        
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Company Status(for efiling)")
                        continue
                elif (Company.columns[i].find("address") != -1):
                    if (Company.columns[i].find("registered") != -1):
                        company_new.insert(index_value, "Registered Address", Company[Company.columns[i]].tolist())
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Registered Address")
                        continue
                elif (Company.columns[i].find("email") != -1):
                    company_new.insert(index_value, "Email Id", Company[Company.columns[i]].tolist())
                    input_file.update({Company.columns[i]: 1})
                    out_file.append("Email Id")
                    continue
                elif (Company.columns[i].find("whether") != -1):
                    if (Company.columns[i].find("listed") != -1):
                        company_new.insert(index_value, "Whether Listed or not",Company[Company.columns[i]].tolist())
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Whether Listed or not")
                        continue
                elif (Company.columns[i].find("par") != -1):
                    if (Company.columns[i].find("value") != -1):
                        company_new.insert(index_value, "Par Value", Company[Company.columns[i]].tolist())
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Par Value")
                        continue
                elif (Company.columns[i].find("address") != -1):
                    if (Company.columns[i].find("registered") == -1):
                        company_new.insert(index_value,"Address other than R/o where all or any books of account and papers are maintained", Company[Company.columns[i]].tolist())
                        input_file.update({Company.columns[i]: 1})
                        out_file.append("Address other than R/o where all or any booCompany.columnss of account and papers are maintained")
                    continue
                elif (Company.columns[i].find("dipp") != -1):
                    company_new.insert(index_value, "Dipp certificate number", Company[Company.columns[i]].tolist())
                    input_file.update({Company.columns[i]: 1})
                    out_file.append("Dipp certificate number")
                    continue
                elif (Company.columns[i].find("industry") != -1):
                    company_new.insert(index_value, "Industry", Company[Company.columns[i]].tolist())
                    input_file.update({Company.columns[i]: 1})
                    out_file.append("Industry")
                    continue
                elif (Company.columns[i].find("website") != -1):
                    company_new.insert(index_value, "Company Website", Company[Company.columns[i]].tolist())
                    input_file.update({Company.columns[i]: 1})
                    out_file.append("Company Website")
                    continue
                elif (Company.columns[i].find("nic") != -1):
                    company_new.insert(index_value, "Nic code", Company[Company.columns[i]].tolist())
                    input_file.update({Company.columns[i]: 1})
                    out_file.append("Nic code")
                    continue
                elif (Company.columns[i].find("business") != -1):
                    company_new.insert(index_value, "Business Details Name", Company[Company.columns[i]].tolist())
                    input_file.update({Company.columns[i]: 1})
                    out_file.append("Business Details Name")
                    continue
            company_new['Date of Incorporation']=pd.to_datetime(company_new['Date of Incorporation'],errors='coerce').dt.strftime('%d-%m-%Y')
            for i in range(len(company_new)):
                if type(company_new['Date of Incorporation'][i]) == str:
                    company_new['Date of Incorporation'][i] = datetime.strptime(str(company_new['Date of Incorporation'][i]), '%d-%m-%Y').date()
            Comparing = ['CIN', 'Company Name', 'Company Category', 'Class of Company',                         'Authorised Capital (INR) as of date', 'Authorised Capital at Incorporation',                         'Paid up Capital (INR)', 'Date of Incorporation', 'Registered Address',                         'Address other than R/o where all or any books of account and papers are maintained',                         'Email Id', 'Whether Listed or not', 'Company Status(for efiling)', 'Dipp certificate number',                         'Par Value', 'Business Details Name', 'Industry', 'Nic code', 'Company Website']
            missing = []
            for x in Comparing:
                if x not in company_new:
                    missing.append(x)
            if len(missing) > 0:
                for i in range(len(missing)):
                    company_new.insert(index_value, missing[i], np.nan)
            out_file.clear()
            Comparing.clear()
            missing.clear()
            continue
        elif (ky[i].find("founder") != -1):
            Founder = pd.read_excel(location, sheet_name=ky.index(ky[i]))
            if Founder.columns.tolist()[1] == 'Unnamed: 1':
                Founder = pd.read_excel(location, sheet_name=ky.index(ky[i]), skiprows=1)
            Founder.drop(Founder.columns[Founder.columns.str.contains('Unnamed', case=False)], axis=1, inplace=True)
            founder_new = pd.DataFrame()
            Founder.columns = map(str.lower, Founder.columns)
            index_value = 0
            out_file = []
            input_file = {k: 0 for v, k in enumerate(Founder.columns)}
            for i in range(len(Founder.columns)):
                if (Founder.columns[i].find("name") != -1):
                    founder_new.insert(index_value, "Name", Founder[Founder.columns[i]].tolist())

                    input_file.update({Founder.columns[i]: 1})
                    out_file.append("Name")
                    continue
                elif (Founder.columns[i].find("nationality") != -1):
                    founder_new.insert(index_value, "Nationality", Founder[Founder.columns[i]].tolist())
                    
                    input_file.update({Founder.columns[i]: 1})
                    out_file.append("Nationality")
                    continue
                elif (Founder.columns[i].find("designation") != -1):
                    founder_new.insert(index_value, "Designation",  Founder[Founder.columns[i]].tolist())
                    
                    input_file.update({Founder.columns[i]: 1})
                    out_file.append("Designation")
                    continue
            nonmand = []
            for key, val in input_file.items():
                if val == 0:
                    nonmand.append(key)
            for i in range(len(nonmand)):
                if (nonmand[i].find("address") != -1):
                    founder_new.insert(index_value,  "Address",  Founder[nonmand[i]].tolist())
                    
                    input_file.update({Founder.columns[i]: 1})
                    out_file.append(
                        "Address")
                    continue
                elif (nonmand[i].find("email") != -1):
                    founder_new.insert(index_value, "Email", Founder[nonmand[i]].tolist())
                    
                    input_file.update({Founder.columns[i]: 1})
                    out_file.append("Email")
                    continue
                elif (nonmand[i].find("contact") != -1):
                    founder_new.insert(index_value, "Contact Number", Founder[nonmand[i]].tolist())
                    
                    input_file.update({Founder.columns[i]: 1})
                    out_file.append("Contact Number")
                    continue
                elif (nonmand[i].find("dob") != -1):
                    founder_new.insert(index_value, "Dob", Founder[nonmand[i]].tolist())
                    
                    input_file.update({Founder.columns[i]: 1})
                    out_file.append("Dob")
                    continue
            Comparing = ['Name', 'Email', 'Nationality', 'Designation', 'Contact Number', 'Address', 'Dob']
            missing = []
            for x in Comparing:
                if x not in founder_new:
                    missing.append(x)
            if len(missing) > 0:
                for i in range(len(missing)):
                    founder_new.insert(index_value, missing[i], np.nan)
                    
            for i in range(len(founder_new)):
                if not founder_new['Nationality'].isnull().any():
                    if (founder_new['Nationality'][i].lower().find("in")) != -1:
                        founder_new['Nationality'][i] = 'Indian'
            founder_new['Dob'] = pd.to_datetime(founder_new['Dob'], errors='coerce').dt.strftime('%d-%m-%Y')
            for i in range(len(founder_new)):
                    if type(founder_new['Dob'][i])==str:
                        founder_new['Dob'][i] = datetime.strptime(str(founder_new['Dob'][i]),'%d-%m-%Y').date()
            founder_mandatory = ['Name', 'Nationality']
            out_file.clear()
            nonmand.clear()
            Comparing.clear()
            missing.clear()
            continue
        elif ky[i].find("director") != -1:
            Director = pd.read_excel(location, sheet_name=ky.index(ky[i]))
            if Director.columns.tolist()[1] == 'Unnamed: 1':
                Director = pd.read_excel(location, sheet_name=ky.index(ky[i]), skiprows=1)
            Director.drop(Director.columns[Director.columns.str.contains('Unnamed', case=False)], axis=1, inplace=True)
            director_new = pd.DataFrame()
            out_file = []

            Director.columns = map(str.lower, Director.columns)
            index_value = 0
            input_file = {k: 0 for v, k in enumerate(Director.columns)}
            for i in range(len(Director.columns)):
                if Director.columns[i].find("name") != -1:
                    director_new.insert(index_value, "Name", Director[Director.columns[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("Name")
                    continue
                elif (Director.columns[i].find("din") != -1):
                    director_new.insert(index_value, "DIN No",   Director[Director.columns[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("DIN No")
                    continue
                elif (Director.columns[i].find("date") != -1):
                    if (Director.columns[i].find("date of appointment") != -1):
                        director_new.insert(index_value, "Date of Appointment",       Director[Director.columns[i]].tolist())
                        
                        input_file.update({Director.columns[i]: 1})
                        out_file.append("Date of Appointment")
                        continue
            nonmand = []
            for key, val in input_file.items():
                if val == 0:
                    nonmand.append(key)
            for i in range(len(nonmand)):
                if (nonmand[i].find("address") != -1):
                    director_new.insert(index_value,   "Address",   Director[nonmand[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append(
                        "Address")
                    continue
                elif (nonmand[i].find("resolution") != -1):
                    director_new.insert(index_value, "Date of board resolution for appointment",   Director[nonmand[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("Date of board resolution for appointment")
                    continue
                elif (nonmand[i].find("residential") != -1):
                    director_new.insert(index_value, "Residential status", Director[nonmand[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("Residential status")
                    continue
                elif (nonmand[i].find("director") != -1):
                    director_new.insert(index_value, "Type of Director", Director[nonmand[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("Type of Director")
                    continue
                elif (nonmand[i].find("resignation") != -1):
                    director_new.insert(index_value, "Date of resignation", Director[nonmand[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("Date of resignation")
                    continue
                elif (nonmand[i].find("email") != -1):
                    director_new.insert(index_value, "Email", Director[nonmand[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("Email")
                    continue
                elif (nonmand[i].find("nationality") != -1):
                    director_new.insert(index_value, "Nationality", Director[nonmand[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("Nationality'")
                    continue
                elif (nonmand[i].find("birth") != -1):
                    director_new.insert(index_value, "dateOfBirth", Director[nonmand[i]].tolist())
                    
                    input_file.update({Director.columns[i]: 1})
                    out_file.append("dateOfBirth")
                    continue
            Comparing = ['DIN No', 'Name', 'Date of Appointment',                         'Date of board resolution for appointment', 'Residential status',                         'Type of Director', 'Date of resignation', 'Email', 'Nationality',                         'dateOfBirth', 'Address']
            missing = []
            for x in Comparing:
                if x not in director_new:
                    missing.append(x)
            if len(missing) > 0:
                for i in range(len(missing)):
                    director_new.insert(index_value, missing[i], np.nan)
                    
            for i in range(len(director_new)):
                if not director_new['Nationality'].isnull().any():
                    if (director_new['Nationality'][i].lower().find("in")) != -1:
                        director_new['Nationality'][i] = 'Indian'
            director_new['Date of Appointment'] = pd.to_datetime(director_new['Date of Appointment'],errors='coerce').dt.strftime('%d-%m-%Y')
            director_new['Date of board resolution for appointment'] = pd.to_datetime(director_new['Date of board resolution for appointment'], errors='coerce').dt.strftime('%d-%m-%Y')
            director_new['Date of resignation'] = pd.to_datetime(director_new['Date of resignation'],errors='coerce').dt.strftime('%d-%m-%Y')
            director_new['dateOfBirth'] = pd.to_datetime(director_new['dateOfBirth'],errors='coerce').dt.strftime('%d-%m-%Y')
            for i in range(len(director_new)):
                if type(director_new['Date of Appointment'][i])==str:
                    director_new['Date of Appointment'][i] = datetime.strptime(str(director_new['Date of Appointment'][i]),'%d-%m-%Y').date()
                if type(director_new['Date of board resolution for appointment'][i])==str:
                    director_new['Date of board resolution for appointment'][i] = datetime.strptime(str(director_new['Date of board resolution for appointment'][i]),'%d-%m-%Y').date()
                if type(director_new['Date of resignation'][i])==str:
                    director_new['Date of resignation'][i] = datetime.strptime(str(director_new['Date of resignation'][i]),'%d-%m-%Y').date()
                if type(director_new['dateOfBirth'][i])==str:
                    director_new['dateOfBirth'][i] = datetime.strptime(str(director_new['dateOfBirth'][i]),'%d-%m-%Y').date()
            out_file.clear()
            nonmand.clear()
            Comparing.clear()
            missing.clear()
            director_mandatory = ['DIN No', 'Name', 'Date of Appointment']
            continue
        elif ky[i].find("first") != -1:
            First_share_holders = pd.read_excel(location, sheet_name=ky.index(ky[i]))
            if 'Name' not in First_share_holders.columns:
                First_share_holders = pd.read_excel(location, sheet_name=ky.index(ky[i]), skiprows=1)
            if 'Unnamed: 3' in First_share_holders.columns:
                First_share_holders.rename(columns={'Unnamed: 3': 'ID Type'}, inplace=True)
            First_share_holders.drop(
                First_share_holders.columns[First_share_holders.columns.str.contains('Unnamed', case=False)], axis=1,                inplace=True)
            fsh_new = pd.DataFrame()
            out_file = []

            First_share_holders.columns = map(str.lower, First_share_holders.columns)
            index_value = 0
            input_file = {k: 0 for v, k in enumerate(First_share_holders.columns)}
            for key, val in input_file.items():
                if val == 0:
                    fsh_new.insert(index_value, key, First_share_holders[key].tolist())
            nonmand = []
            for key, val in input_file.items():
                if val == 0:
                    nonmand.append(key)
            for i in range(len(nonmand)):
                if (nonmand[i].find("name") != -1):
                    fsh_new.insert(index_value, "Name", First_share_holders[nonmand[i]].tolist())
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Name")
                    continue
                if (nonmand[i].find("share") != -1):
                    fsh_new.insert(index_value, "No. of shares", First_share_holders[nonmand[i]].tolist())
                    input_file.update({nonmand[i]: 1})
                    out_file.append("No. of shares")
                    continue
                if (nonmand[i].find("pan") != -1):
                    fsh_new.insert(index_value, "ID #", First_share_holders[nonmand[i]].tolist())
                    input_file.update({nonmand[i]: 1})
                    out_file.append("ID #")
                    continue
            if nonmand[i].find("id") != -1:
                if (nonmand[i].find("no") != -1):
                    fsh_new.insert(index_value, "ID #", First_share_holders[nonmand[i]].tolist())
                    input_file.update({nonmand[i]: 1})
                    out_file.append("ID #")
                    continue
                if (nonmand[i].find("type") != -1):
                    fsh_new.insert(index_value, "ID type", First_share_holders[nonmand[i]].tolist())
                    input_file.update({nonmand[i]: 1})
                    out_file.append("ID type")
                    continue
            fsh_new.dropna(subset=['No. of shares'], inplace=True)
            fsh_new.reset_index(inplace=True)
            continue
        elif (ky[i].find("round") != -1):
            Round_Creation = pd.read_excel(location, sheet_name=ky.index(ky[i]))
            if Round_Creation.columns.tolist()[1] == 'Unnamed: 1':
                Round_Creation = pd.read_excel(location, sheet_name=ky.index(ky[i]), skiprows=1)
            round_new = pd.DataFrame()
            out_file = []
            Round_Creation.columns = map(str.lower, Round_Creation.columns)
            index_value = 0
            input_file = {k: 0 for v, k in enumerate(Round_Creation.columns)}
            Round_Creation.rename(columns={'pan no. of shareholders': 'Id#'}, inplace=True)
            for i in range(len(Round_Creation.columns)):
                if (Round_Creation.columns[i].find("share") != -1):
                    if (Round_Creation.columns[i].find("name") != -1):
                        if (Round_Creation.columns[i].find('name of shareholder') != -1):
                            if (Round_Creation.columns[i].find('legal') == -1):
                                round_new.insert(index_value, "Name of shareholder",Round_Creation[Round_Creation.columns[i]].tolist())
                                
                                input_file.update({Round_Creation.columns[i]: 1})
                                out_file.append("Name of shareholder")
                                continue
                    elif (Round_Creation.columns[i].find("type") != -1):
                        if (Round_Creation.columns[i].find("shares") != -1):
                            if (Round_Creation.columns[i].find("other") == -1):
                                round_new.insert(index_value, "Type of shares",            Round_Creation[Round_Creation.columns[i]].tolist())
                                
                                input_file.update({Round_Creation.columns[i]: 1})
                                out_file.append("Type of shares")
                                continue
                        elif (Round_Creation.columns[i].find("shareholders") != -1):
                            round_new.insert(index_value, "Type of Shareholder",        Round_Creation[Round_Creation.columns[i]].tolist())
                            
                            input_file.update({Round_Creation.columns[i]: 1})
                            out_file.append("Type of Shareholder")
                            continue
                    elif (Round_Creation.columns[i].find("per") != -1):
                        if (Round_Creation.columns[i].find("par") != -1):
                            round_new.insert(index_value, "Par value per share",        Round_Creation[Round_Creation.columns[i]].tolist())
                            
                            input_file.update({Round_Creation.columns[i]: 1})
                            out_file.append("Par value per share")
                            continue
                        elif (Round_Creation.columns[i].find("premium") != -1):
                            round_new.insert(index_value, "Premium Per share",        Round_Creation[Round_Creation.columns[i]].tolist())
                            
                            input_file.update({Round_Creation.columns[i]: 1})
                            out_file.append("Premium Per share")
                            continue
                        elif (Round_Creation.columns[i].find("price") != -1):
                            if (Round_Creation.columns[i].find("applicable") == -1):
                                round_new.insert(index_value, "Per share price",            Round_Creation[Round_Creation.columns[i]].tolist())
                                
                                input_file.update({Round_Creation.columns[i]: 1})
                                out_file.append("Per share price")
                                continue
                    elif (Round_Creation.columns[i].find("number") != -1) or (
                            Round_Creation.columns[i].find("no") != -1):
                        if (Round_Creation.columns[i].find("shares") != -1):
                            round_new.insert(index_value, "No. of shares",        Round_Creation[Round_Creation.columns[i]].tolist())
                            
                            input_file.update({Round_Creation.columns[i]: 1})
                            out_file.append("No. of shares")
                            continue
                if (Round_Creation.columns[i].find("round") != -1):
                    if (Round_Creation.columns[i].find("name") != -1):
                        round_new.insert(index_value, "Round Name",    Round_Creation[Round_Creation.columns[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append("Round Name")
                        continue
                elif (Round_Creation.columns[i].find("allotment") != -1):
                    if (Round_Creation.columns[i].find("resolution") == -1):
                        round_new.insert(index_value, "Date of allotment",    Round_Creation[Round_Creation.columns[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append("Date of allotment")
                        continue
                elif (Round_Creation.columns[i].find("total") != -1):
                    round_new.insert(index_value, "Total amount invested",Round_Creation[Round_Creation.columns[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Total amount invested")
                    continue
                elif (Round_Creation.columns[i].find("valuation") != -1):
                    if (Round_Creation.columns[i].find("pre") != -1):
                        round_new.insert(index_value, "Conversion", Round_Creation[Round_Creation.columns[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append("Conversion")
                        continue
            nonmand = []
            for key, val in input_file.items():
                if val == 0:
                    nonmand.append(key)
            for i in range(len(nonmand)):
                if (nonmand[i].find("form of consideration") != -1):
                    round_new.insert(index_value, "Form of consideration", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Form of consideration")
                    continue
                if (nonmand[i].find("nationality") != -1):
                    round_new.insert(index_value,"Nationality",Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append(
                        "Nationality")
                    continue
                if (nonmand[i].find("id") != -1):
                    if (nonmand[i].find("#") != -1):
                        round_new.insert(index_value,    "ID#",    Round_Creation[nonmand[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append(
                            "ID#")
                        continue
                    elif (nonmand[i].find("type") != -1):
                        round_new.insert(index_value, "ID type", Round_Creation[nonmand[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append("ID type")
                        continue
                if (nonmand[i].find("nationlity") != -1):
                    round_new.insert(index_value, "Nationality", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Nationality")
                    continue
                if (nonmand[i].find("residential") != -1):
                    round_new.insert(index_value, "Residential status", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Residential status")
                    continue
                if (nonmand[i].find("fx") != -1):
                    round_new.insert(index_value, "Per share price (FX) if applicable",Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Per share price (FX) if applicable")
                    continue
                if (nonmand[i].find("distinctive") != -1):
                    round_new.insert(index_value, "Share distinctive numbers", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Share distinctive numbers")
                    continue
                if (nonmand[i].find("consideration") != -1):
                    round_new.insert(index_value, "Form of consideration", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Form of consideration")
                    continue
                if (nonmand[i].find("certificate") != -1):
                    round_new.insert(index_value, "Share certificate number", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Share certificate number")
                    continue
                if (nonmand[i].find("folio") != -1):
                    round_new.insert(index_value, "Folio Number", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Folio Number")
                    continue
                if (nonmand[i].find("relevant") != -1):
                    round_new.insert(index_value, "Relevant Shareholders Agreement",Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Relevant Shareholders Agreement")
                    continue
                if (nonmand[i].find("reference") != -1):
                    round_new.insert(index_value, "Reference", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Reference")
                    continue
                if (nonmand[i].find("report") != -1):
                    round_new.insert(index_value, "Valuation as per Report(Pas 4)", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Valuation as per Report(Pas 4)")
                    continue
                if (nonmand[i].find("lead") != -1):
                    round_new.insert(index_value, "Lead Investor", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Lead Investor")
                    continue
                if (nonmand[i].find("comments") != -1):
                    round_new.insert(index_value, "Comments", Round_Creation[nonmand[i]].tolist())
                    
                    input_file.update({Round_Creation.columns[i]: 1})
                    out_file.append("Comments")
                    continue
                if (nonmand[i].find("type of") != -1):
                    if (nonmand[i].find("shareholder") != -1):
                        round_new.insert(index_value, "Type of Shareholder", Round_Creation[nonmand[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append("Type of Shareholder")
                        continue
                if (nonmand[i].find("name") != -1):
                    if (nonmand[i].find("legal") != -1):
                        round_new.insert(index_value, "Legal Name of shareholder", Round_Creation[nonmand[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append("Legal Name of shareholder")
                        continue
                    elif (nonmand[i].find("group") != -1):
                        round_new.insert(index_value, "Group Name", Round_Creation[nonmand[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append("Group Name")
                        continue
                    elif (nonmand[i].find("instrument") != -1):
                        round_new.insert(index_value, "Instrument Name or Security Name",    Round_Creation[nonmand[i]].tolist())
                        
                        input_file.update({Round_Creation.columns[i]: 1})
                        out_file.append("Instrument Name or Security Name")
                        continue
            for i in range(len(round_new.columns)):
                if Round_Creation.columns[i].find("unnamed") != -1:
                    log = log.append({'Sheet Name': 'Round Creation', 'Error At': Round_Creation.columns[i]},ignore_index=True)

                Comparing = ['Name of shareholder', 'No. of shares', 'ID#', 'ID type',                             'Type of shares', 'Type of Shareholder', 'Par value per share',                             'Round Name', 'Date of allotment', 'Legal Name of shareholder',                             'Nationality', 'Residential status', 'Instrument Name or Security Name',                             'Premium Per share', 'Per share price',                             'Per share price (FX) if applicable', 'Total amount invested',                             'Form of consideration', 'Conversion', 'Share distinctive numbers',                             'Share certificate number', 'Folio Number',                             'Relevant Shareholders Agreement', 'Reference',                             'Valuation as per Report(Pas 4)', 'Group Name', 'Lead Investor',                             'Comments']
            missing = []
            for x in Comparing:
                if x not in round_new:
                    missing.append(x)
            if len(missing) > 0:
                for i in range(len(missing)):
                    round_new.insert(index_value, missing[i], np.nan)
                    
            out_file.clear()
            nonmand.clear()
            Comparing.clear()
            missing.clear()
            round_mandatory = ['Name of shareholder', 'No. of shares',                               'Type of shares', 'Par value per share',                               'Round Name', 'Date of allotment',                               'Per share price',                               'Total amount invested']
            continue
        elif (ky[i].find("secondary") != -1):
            Secondary = pd.read_excel(location, sheet_name=ky.index(ky[i]))
            if Secondary.columns.tolist()[1] == 'Unnamed: 1':
                Secondary = pd.read_excel(location, sheet_name=ky.index(ky[i]), skiprows=1)
            Secondary.drop(Secondary.columns[Secondary.columns.str.contains('Unnamed', case=False)], axis=1,                           inplace=True)
            secondary_new = pd.DataFrame()
            out_file = []

            Secondary.rename(columns={'Instrument': 'Type of shares'}, inplace=True)
            Secondary.columns = map(str.lower, Secondary.columns)
            index_value = 0
            input_file = {k: 0 for v, k in enumerate(Secondary.columns)}
            nonmand = []
            for key, val in input_file.items():
                if val == 0:
                    nonmand.append(key)
            for i in range(len(nonmand)):
                if (nonmand[i].find("seller") != -1):
                    secondary_new.insert(index_value, "Seller name", Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Seller name")
                    continue
                if (nonmand[i].find("date") != -1):
                    secondary_new.insert(index_value, "date of Transfer", Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("date of Transfer")
                    continue
                if (nonmand[i].find("share") != -1):
                    if (nonmand[i].find("type of") != -1):
                        secondary_new.insert(index_value, "Type of Shares", Secondary[nonmand[i]].tolist())
                        
                        input_file.update({nonmand[i]: 1})
                        out_file.append("Type of Shares")
                        continue
                    elif (nonmand[i].find("number of") != -1)or (nonmand[i].find("no. of") != -1):
                        if (nonmand[i].find("total") == -1):
                            secondary_new.insert(index_value, "Number of shares", Secondary[nonmand[i]].tolist())
                            
                            input_file.update({nonmand[i]: 1})
                            out_file.append("Number of shares")
                            continue
                    elif (nonmand[i].find("price") != -1):
                        secondary_new.insert(index_value, "Price per share", Secondary[nonmand[i]].tolist())
                        
                        input_file.update({nonmand[i]: 1})
                        out_file.append("Price per share")
                        continue
                    elif (nonmand[i].find("start") != -1):
                        secondary_new.insert(index_value, "Share start number", Secondary[nonmand[i]].tolist())
                        
                        input_file.update({nonmand[i]: 1})
                        out_file.append("Share start number")
                        continue
                    elif (nonmand[i].find("end") != -1):
                        secondary_new.insert(index_value, "Share end number", Secondary[nonmand[i]].tolist())
                        
                        input_file.update({nonmand[i]: 1})
                        out_file.append("Share end number")
                        continue
                    elif (nonmand[i].find("holder") != -1):
                        secondary_new.insert(index_value, "Shareholder Type", Secondary[nonmand[i]].tolist())
                        
                        input_file.update({nonmand[i]: 1})
                        out_file.append("Shareholder Type")
                        continue
                if (nonmand[i].find("total") != -1):
                    if (nonmand[i].find("price") != -1):
                        secondary_new.insert(index_value, "total Price", Secondary[nonmand[i]].tolist())
                        
                        input_file.update({nonmand[i]: 1})
                        out_file.append("total Price")
                        continue
                if (nonmand[i].find("buyer") != -1):
                    if (nonmand[i].find("group") == -1):
                        secondary_new.insert(index_value, "Buyer Name", Secondary[nonmand[i]].tolist())
                        input_file.update({nonmand[i]: 1})
                        out_file.append("Buyer Name")
                        continue
                    else:
                        secondary_new.insert(index_value, "Group Buyer Name", Secondary[nonmand[i]].tolist())
                        
                        input_file.update({nonmand[i]: 1})
                        out_file.append("Group Buyer Name")
                        continue
                if (nonmand[i].find("instrument") != -1):
                    secondary_new.insert(index_value, "Instrument Name or Security Name",    Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Instrument Name or Security Name")
                    continue
                if (nonmand[i].find("address") != -1):
                    secondary_new.insert(index_value, "Registerd Address", Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Registerd Address")
                    continue
                if (nonmand[i].find("status") != -1):
                    secondary_new.insert(index_value, "Resident Status", Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Resident Status")
                    continue
                if (nonmand[i].find("occupation") != -1):
                    secondary_new.insert(index_value, "Occupation", Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Occupation")
                    continue
                if (nonmand[i].find("nationality") != -1):
                    secondary_new.insert(index_value, "Nationality", Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Nationality")
                    continue
                if (nonmand[i].find("number") != -1):
                    if (nonmand[i].find("contact") != -1):
                        secondary_new.insert(index_value, "Contact Number", Secondary[nonmand[i]].tolist())
                        
                        input_file.update({nonmand[i]: 1})
                        out_file.append("Contact Number")
                    continue
                if (nonmand[i].find("email") != -1):
                    secondary_new.insert(index_value, "email Id", Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("email Id")
                    continue
                if (nonmand[i].find("comment") != -1):
                    secondary_new.insert(index_value, "Comment", Secondary[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Comment")
                    continue
            Comparing = ['Seller name', 'ID #', 'ID Type', 'date of Transfer', 'Type of Shares', 'Number of shares',                         'Instrument Name or Security Name', 'Price per share', 'total Price', 'Share start number',                         'Share end number', 'Buyer Name', 'Group Name Buyer', 'Shareholder Type', 'Registerd Address',                         'Resident Status', 'Occupation', 'Nationality', 'Contact Number', 'email Id', 'Comment']
            Sec_mandatory = ['Seller name', 'date of Transfer', 'Type of Shares', 'Number of shares', 'Price per share',                             'total Price', 'Buyer Name']
            missing = []
            for x in Comparing:
                if x not in secondary_new:
                    missing.append(x)
            if len(missing) > 0:
                for i in range(len(missing)):
                    secondary_new.insert(index_value, missing[i], np.nan)
            out_file.clear()
            nonmand.clear()
            Comparing.clear()
            missing.clear()
            continue
        elif (ky[i].find("esop") != -1):
            Esop = pd.read_excel(location, sheet_name=ky.index(ky[i]))
            if Esop.columns.tolist()[1] == 'Unnamed: 1':
                Esop = pd.read_excel(location, sheet_name=ky.index(ky[i]), skiprows=1)
            Esop.drop(Esop.columns[Esop.columns.str.contains('Unnamed', case=False)], axis=1, inplace=True)
            esop_new = pd.DataFrame()
            out_file = []

            Esop.columns = map(str.lower, Esop.columns)
            index_value = 0
            input_file = {k: 0 for v, k in enumerate(Esop.columns)}
            for key, val in input_file.items():
                if val == 0:
                    esop_new.insert(index_value, key, Esop[key].tolist())
            nonmand = []
            for key, val in input_file.items():
                if val == 0:
                    nonmand.append(key)
            Esop_mandatory = ['financial year', 'pool']
            for i in range(len(nonmand)):
                if (nonmand[i].find("year") != -1):
                    esop_new.insert(index_value, "Financial Year", Esop[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Financial Year")
                    continue
                if (nonmand[i].find("pool") != -1):
                    esop_new.insert(index_value, "Pool", Esop[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Pool")
                    continue
                if (nonmand[i].find("granted") != -1):
                    esop_new.insert(index_value, "Granted", Esop[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Granted")
                    continue
                if (nonmand[i].find("vested") != -1):
                    esop_new.insert(index_value, "Vested", Esop[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Vested")
                    continue
                if (nonmand[i].find("exercised") != -1):
                    esop_new.insert(index_value, "Exercised", Esop[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Exercised")
                    continue
                if (nonmand[i].find("forfeited") != -1):
                    esop_new.insert(index_value, "Forfeited", Esop[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Forfeited")
                    continue
                if (nonmand[i].find("available") != -1):
                    esop_new.insert(index_value, "Available for Grant", Esop[nonmand[i]].tolist())
                    
                    input_file.update({nonmand[i]: 1})
                    out_file.append("Available for Grant")
                    continue
            Comparing = ['Available for Grant', 'Forfeited', 'Exercised', 'Vested', 'Granted', 'Pool', 'Financial Year']
            missing = []
            for x in Comparing:
                if x not in esop_new:
                    missing.append(x)
            if len(missing) > 0:
                for i in range(len(missing)):
                    esop_new.insert(index_value, missing[i], np.nan)
            out_file.clear()
            nonmand.clear()
            Comparing.clear()
            missing.clear()
            continue
    if 'esop' not in ky:
        print("Creating ESOP")
        esop_new = pd.DataFrame()
        esop_new.insert(0, "Financial Year", np.nan)
        esop_new.insert(1, "Pool", np.nan)
        esop_new.insert(2, "Granted", np.nan)
        esop_new.insert(3, "Vested", np.nan)
        esop_new.insert(4, "Exercised", np.nan)
        esop_new.insert(5, "Forfeited", np.nan)
        esop_new.insert(6, "Available for Grant", np.nan)
    sheet = []
    for i in range(len(wb.sheetnames)):
        sheet.append(wb.sheetnames[i].lower())
    company_new.dropna(subset=['Company Name'], inplace=True)
    company_new.reset_index(inplace=True)
    dictionary = {'Company Name': company_new['Company Name'][0],'CIN':str(company_new['CIN'][0]).strip(),'Status': np.nan,'company Par value missing': 'No','Round Creation Round Name missing': 'No','Round Creation PPS missing': 'No','Conversion of shares is there?': 'No','Split is there?':'No','Partly Paid':'No','Paidup Matched': 'No','Round Creation Less than 50 entries': 'No'
                  ,'Secondary Errors' : 'No','Numberoferrors': 0 ,'Review Status':''}
    if company_new['CIN'][0] in cinlist:
            print(str(company_new['Company Name'][0])+" is already existing")
            return basepath, dictionary
    for i in range(len(sheet)):
        if sheet[i].find("first") != -1:
            Name_of_fsh = fsh_new['Name'].tolist()
            Numberofshares = fsh_new['No. of shares'].tolist()
            if 'ID type' not in fsh_new.columns.tolist():
                fsh_new.insert(3,'ID type',np.nan)
            if 'ID #' not in fsh_new.columns.tolist():
                fsh_new.insert(2,'ID #',np.nan)
            IDType = fsh_new['ID type'].tolist()
            ID = fsh_new['ID #'].tolist()
            Typeofshares = []
            Typeofshareholders = []
            Parvaluepershare = []
            Roundname = []
            Dateofallot = []
            for i in range(0, len(Numberofshares)):
                Typeofshareholders.append('Founders')
                Typeofshares.append('Equity')
                if type(company_new['Par Value'][0]) == int:
                    Parvaluepershare.append(company_new['Par Value'][0])
                else:
                    Parvaluepershare.append(10)
                    log.append({'Sheet Name': 'Company', 'Warning At': 'Par Value','Fixed Legal Team': 'Default value 10 was inserted'}, ignore_index=True)
                Roundname.append('Incorporation Round')
                Dateofallot.append(company_new['Date of Incorporation'][0])
            namerow = pd.DataFrame(
                {'Name of shareholder': Name_of_fsh, 'No. of shares': Numberofshares, 'ID#': ID, 'ID type': IDType,                 'Type of shares': Typeofshares, 'Type of Shareholder': Typeofshareholders,                 'Par value per share': Parvaluepershare, 'Round Name': Roundname, 'Date of allotment': Dateofallot})
            round_new = pd.concat([namerow, round_new]).reset_index(drop=True)
    founder_new.dropna(subset=['Name'], inplace=True)
    founder_new.reset_index(inplace=True)
    director_new.dropna(subset=['Name'], inplace=True)
    director_new.reset_index(inplace=True)
    round_new.dropna(subset=['Par value per share', 'Name of shareholder'], inplace=True)
    round_new.reset_index(inplace=True)
    for i in range(len(round_new)):
        if round_new['Type of shares'][i]=='NCDs':
            round_new['Type of shares'][i]='NCD'
    ans = 0
    company_new['Nic code'] = company_new['Nic code'].astype(str)
    company_new['Authorised Capital (INR) as of date']=company_new['Authorised Capital (INR) as of date'].replace(",", "", regex=True)
    for i in range(len(company_new)):
        if company_new['Authorised Capital (INR) as of date'][i]!=np.nan:
            company_new['Authorised Capital (INR) as of date'][i]=int(company_new['Authorised Capital (INR) as of date'][i])
    round_new['No. of shares'] = round_new['No. of shares'].replace(",", "", regex=True)
    round_new['Par value per share'] = round_new['Par value per share'].replace(",", "", regex=True)
    round_new['Per share price'] = round_new['Per share price'].replace(",", "", regex=True)
    round_new['Premium Per share'] = round_new['Premium Per share'].replace(",","",regex=True)
    round_new['Premium Per share'] = round_new['Premium Per share'].fillna(0)
    for i in range(len(round_new)):
        try:
            round_new['Per share price'][i]=float(round_new['Per share price'][i])
        except:
            round_new['Per share price'][i]=0
    for i in range(len(round_new)):
        try:
            round_new['Premium Per share'][i]=float(round_new['Premium Per share'][i])
        except:
            round_new['Premium Per share'][i]=0
    for i in range(len(round_new)):
        try:
            round_new['No. of shares'][i] = int(round_new['No. of shares'][i])
        except:
            round_new['No. of shares'][i] = round_new['No. of shares'][i]
    for i in range(len(round_new)):
        round_new['Share distinctive numbers'][i] = np.nan
    for i in range(len(round_new)):
        round_new['Share certificate number'][i] = np.nan
    round_new['Par value per share'] = round_new['Par value per share'].fillna(10)
    round_new['Premium Per share'] = round_new['Premium Per share'].astype(float).round(2)
    # Finding Incorporation Round
    for i in range(len(round_new)):
        if round_new['Round Name'].isnull().any() == False:
            if round_new['Round Name'][i].lower().find("incorpo") != -1:
                round_new['Round Name'][i] = 'Incorporation Round'
    # Converting ID# to string
    round_new['ID#'] = round_new['ID#'].astype(str)
    round_new['ID#'] = round_new['ID#'].replace("nan", "", regex=True)
    # Converting name to Title Case
    for i in range(len(founder_new)):
        if type(founder_new['Name'][i]) == str:
            founder_new['Name'][i] = founder_new['Name'][i].title()
    for i in range(len(director_new)):
        if type(director_new['Name'][i]) == str:
            director_new['Name'][i] = director_new['Name'][i].title()
    for i in range(len(round_new)):
        if type(round_new['Name of shareholder'][i]) == str and "fund" not in round_new['Name of shareholder'][i].lower() and "venture" not in round_new['Name of shareholder'][i].lower() and "capital" not in round_new['Name of shareholder'][i].lower() and "investment" not in round_new['Name of shareholder'][i].lower():
            round_new['Name of shareholder'][i] = round_new['Name of shareholder'][i].title()
    round_new['Per share price']=round_new['Par value per share'].astype(float)+round_new['Premium Per share'].astype(float)
    round_new['Per share price'] = round_new['Per share price'].astype(float).round(2)
    round_new['Total amount invested'] = round_new['Per share price'] *round_new['No. of shares']
    round_new['Total amount invested'] = round_new['Total amount invested'].astype(float).round(2)
    # Calculating Authorised Capital at Incorporation
    for i in range(len(round_new)):
        if round_new['Round Name'][i] == 'Incorporation Round':
            ans = ans + float(round_new['Total amount invested'][i])
    for i in range(len(company_new)):
        company_new['Authorised Capital at Incorporation'][i] = (ans)
    for i in range(len(round_new)):
        if round_new['Lead Investor'][i]==0:
            round_new['Lead Investor'][i]='No'
        if round_new['Lead Investor'][i]==1:
            round_new['Lead Investor'][i]='Yes'
        if round_new['Lead Investor'][i]=='True':
            round_new['Lead Investor'][i]='Yes'
        if round_new['Lead Investor'][i]=='False':
            round_new['Lead Investor'][i]='No'
    security_mandatory = ['Round Name', 'Price per share', 'No. of Shares in this round', 'Total Premoney Shares','Pre Money Valuation', 'Post Money Valuation', 'Type of Shares', 'Ratio']
    security_new = pd.DataFrame(
        columns=['Round Name', 'Price per share', 'No. of Shares in this round', 'Total Premoney Shares',                 'Pre Money Valuation', 'Post Money Valuation', 'Instrument Name or Security Name', 'Type of Shares',                 'Par value', 'Voting Rights', 'Dividend type', 'Dividend to equity holders', 'Dividend Rate',                 'Pari Passu', 'Rank', 'Amount', 'Participation', 'Cap Value', 'Ratio', 'Conversion Period',                 'Valuation adjustment based on convertible', 'Conversion Forumula', 'Discount/Conversion terms',                 'Valuation Cap (INR)', 'Anti- dilution', 'Stock restructuring', 'Number of warrants', 'Exercise price',                 'Upfront amount', 'Lock-in period', 'Exercise Period', 'Redemption Amount', 'Note Period',                 'Note Amount', 'Note Interest', 'Note Payment', 'Comment'])
    Roundcreationround = round_new['Round Name'].tolist()
    Roundcreationround = removeduplicate(Roundcreationround)
    if 'Incorporation Round' in Roundcreationround:
        Roundcreationround.remove('Incorporation Round')
    security_new['Round Name'] = Roundcreationround
    Roundname = round_new['Round Name'].replace('Round', '', regex=True)
    Typeofsharess = round_new['Type of shares']
    
    round_new['Instrument Name or Security Name'] = Roundname + Typeofsharess
    for i in range(len(security_new['Round Name'])):
        s = 0
        for j in range(len(round_new['Round Name'])):
            if security_new['Round Name'][i] == round_new['Round Name'][j]:
                security_new['Price per share'][i] = float(round_new['Per share price'][j])
                security_new['Instrument Name or Security Name'][i] = round_new['Instrument Name or Security Name'][j]
                security_new['Type of Shares'][i] = round_new['Type of shares'][j]
                security_new['Par value'][i] = round_new['Par value per share'][j]
                if round_new['Type of shares'][i] == 'Equity' or round_new['Type of shares'][i] == 'CCPS':
                    s = s + int(round_new['No. of shares'][j])
                    security_new['No. of Shares in this round'] = s
    #security_new['Price per share']=security_new['Price per share'].astype(float).round(2)
    # Security Calculation part
    for i in range(len(security_new['Round Name'])):
        sum = 0
        for j in range(len(round_new['Round Name'])):
            if security_new['Round Name'][i] == round_new['Round Name'][j]:
                #security_new['Price per share'][i] = float(round_new['Per share price'][j])
                if round_new['Type of shares'][j] == 'Equity' or round_new['Type of shares'][j] == 'CCPS':
                    sum = sum + int(round_new['No. of shares'][j])
                security_new['No. of Shares in this round'][i] = sum
    # for i in range(len(security_new['Round Name'])):
    #     security_new['Post Money Valuation'][i] = ((security_new['Total Premoney Shares'][i]) + (
    #         security_new['No. of Shares in this round'][i])) * (security_new['Price per share'][i])
    # for i in range(len(security_new)):
    #     for j in range(len(round_new)):
    #         if security_new['Round Name'][i] == round_new['Round Name'][j]:
    #             round_new['Instrument Name or Security Name'][j] = security_new['Instrument Name or Security Name'][i]
    a = 0
    for j in range(len(security_new)):
        if type(security_new['Conversion Period'][j]) == str:
            security_new['Conversion Period'][j] = 20
    amountpaid = []
    for i in range(len(round_new)):
        if round_new['Type of shares'][i] == 'EQUITY' or round_new['Type of shares'][i] == 'Equity' or round_new['Type of shares'][i] == 'CCPS':
            amountpaid.append((round_new['No. of shares'][i]) * (round_new['Par value per share'][i]))
    sum = 0
    for i in range(len(amountpaid)):
        sum = sum + float(amountpaid[i])
    try:
        for i in range(len(secondary_new)):
            if secondary_new['total Price'][i] == np.nan:
                secondary_new['total Price'][i] = secondary_new['Number of shares'][i] * \
             secondary_new['Price per share'][i]
    except:
        secondary_new['total Price'] = secondary_new['total Price']
    try:
        secondary_new['total Price']=secondary_new['Number of shares']*secondary_new['Price per share']
    except:
        secondary_new['total Price']=secondary_new['total Price']
    round_new['Date of allotment'] = pd.to_datetime(round_new['Date of allotment'], errors='coerce').dt.strftime('%d-%m-%Y')
    for i in range(len(round_new)):
        if type(round_new['Date of allotment'][i])==str:
            round_new['Date of allotment'][i] = datetime.strptime(str(round_new['Date of allotment'][i]), '%d-%m-%Y').date()
    totalsum = 0
    for i in range(len(round_new['Round Name'])):
        if round_new['Round Name'][i] == 'Incorporation Round':
            totalsum = totalsum + round_new['No. of shares'][i]
    for i in range(len(security_new['Round Name'])):
        if i >= 2:
            if security_new['No. of Shares in this round'][i] == 0 and security_new['Price per share'][i] == \
                    security_new['Price per share'][i - 1] and len(security_new) != 1:
                security_new['Total Premoney Shares'][i] = security_new['Total Premoney Shares'][i - 1]
                totalsum = totalsum + security_new['No. of Shares in this round'][i]
                security_new['No. of Shares in this round'][i] = security_new['No. of Shares in this round'][i - 1]
            else:
                security_new['Total Premoney Shares'][i] = totalsum
                totalsum = totalsum + security_new['No. of Shares in this round'][i]
        else:
            security_new['Total Premoney Shares'][i] = totalsum
            totalsum = totalsum + security_new['No. of Shares in this round'][i]
    security_new['Pre Money Valuation'] = security_new['Price per share'] * security_new['Total Premoney Shares']
    for i in range(len(security_new['Round Name'])):
        security_new['Post Money Valuation'][i] = ((security_new['Total Premoney Shares'][i]) + (
            security_new['No. of Shares in this round'][i])) * (security_new['Price per share'][i])
    for i in range(len(round_new)):
        for j in range(i + 1, len(round_new)):
            if round_new['Name of shareholder'][i] == round_new['Name of shareholder'][j] and round_new['Round Name'][
                i] == round_new['Round Name'][j]:
                if round_new['Type of shares'][i] == 'Equity' and round_new['Type of shares'][j] == 'CCPS':
                    round_new['Type of shares'][i] = 'Nominal'
                elif round_new['Type of shares'][i] == 'CCPS' and round_new['Type of shares'][j] == 'Equity':
                    round_new['Type of shares'][j] = 'Nominal'
    round_new.sort_values(['Date of allotment', 'Round Name', 'Name of shareholder'], ascending=[True, True, True],                          inplace=True)
    for i in range(len(security_new)):
        if type(security_new['Dividend Rate'][i]) == str:
            security_new['Dividend Rate'] = security_new['Dividend Rate'].apply(lambda x: find_number(x))
    for i in range(len(security_new)):
        if type(security_new['Amount'][i]) == str and security_new['Amount'][i] != np.nan:
            security_new['Amount'][i] = find_number(security_new['Amount'][i])
    for i in range(len(security_new)):
        if security_new['Valuation adjustment based on convertible'][i] != np.nan:
            security_new['Valuation adjustment based on convertible'][i] = np.nan
    for i in range(len(security_new)):
        if security_new['Conversion Forumula'][i] != np.nan:
            security_new['Conversion Forumula'][i] = np.nan
    for i in range(len(security_new)):
        if security_new['Discount/Conversion terms'][i] != np.nan:
            security_new['Discount/Conversion terms'][i] = np.nan
    for i in range(len(security_new)):
        if security_new['Valuation Cap (INR)'][i] != np.nan:
            security_new['Valuation Cap (INR)'][i] = np.nan
    for i in range(len(security_new)):
        if type(security_new['Cap Value'][i]) == str:
            security_new['Cap Value'][i] = 0
    for i in range(len(round_new)):
        if round_new['ID#'][i] == np.nan and round_new['Round Name'][i] == 'Incorporation Round':
            round_new['ID#'][i] = 'ABCDEF1234G'
            round_new['ID type'][i] = 'Others'
    round_new['Conversion']='Fixed'
    for i in range(len(secondary_new)):
        secondary_new['Instrument Name or Security Name'][i] = secondary_new['Type of Shares'][i]
    company_mandatory = ['CIN', 'Company Name', 'Company Category', 'Class of Company',                         'Authorised Capital (INR) as of date', 'Authorised Capital at Incorporation',                         'Paid up Capital (INR)', 'Date of Incorporation', 'Registered Address', 'Email Id',                         'Whether Listed or not', 'Company Status(for efiling)', 'Par Value']
    if company_new['Company Category'].isnull().any():
        company_new['Company Category'] = company_new['Company Category'].fillna('Company limited by Shares')
        log = log.append({'Sheet Name': 'Company', 'Warning At': 'Company Category',                          'Fixed Legal Team': 'Default value Company limited by Shares was inserted'},                         ignore_index=True)
    if company_new['Class of Company'].isnull().any():
        company_new['Class of Company'] = company_new['Class of Company'].fillna('Private Limited')
        log = log.append({'Sheet Name': 'Company', 'Warning At': 'Class of Company',                          'Fixed Legal Team': 'Default value Private Limited was inserted'}, ignore_index=True)
    if company_new['Company Status(for efiling)'].isnull().any():
        company_new['Company Status(for efiling)'] = company_new['Company Status(for efiling)'].fillna('Active')
        log = log.append({'Sheet Name': 'Company', 'Warning At': 'Company Status(for efiling)',                          'Fixed Legal Team': 'Default value Active was inserted'}, ignore_index=True)
    if director_new['Type of Director'].isnull().any():
        director_new['Type of Director'] = director_new['Type of Director'].fillna('Executive')
        log = log.append({'Sheet Name': 'Director', 'Warning At': 'Type of Director',                          'Fixed Legal Team': 'Default value Executive was inserted'}, ignore_index=True)
    if round_new['Type of shares'].isnull().any():
        round_new['Type of shares'] = round_new['Type of shares'].fillna('Others')
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'Type of shares',                          'Fixed Legal Team': 'Default value Others was inserted'}, ignore_index=True)
    if round_new['Premium Per share'].isnull().any():
        round_new['Premium Per share'] = round_new['Premium Per share'].fillna(0)
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'Premium Per share',                          'Fixed Legal Team': 'Default value 0 was inserted'}, ignore_index=True)
    if round_new['Nationality'].isnull().any():
        round_new['Nationality'] = round_new['Nationality'].fillna('Foreign')
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'Nationality',                          'Fixed Legal Team': 'Default value Foreign was inserted'}, ignore_index=True)
    if round_new['Residential status'].isnull().any():
        round_new['Residential status'] = round_new['Residential status'].fillna('Non-Resident')
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'Residential status',                          'Fixed Legal Team': 'Default value Non-Resident was inserted'}, ignore_index=True)
    if round_new['Total amount invested'].isnull().any():
        round_new['Total amount invested'] = round_new['Total amount invested'].fillna(
            (round_new['No. of shares']) * (round_new['Par value per share']))
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'Total amount invested',                          'Fixed Legal Team': 'Calculated value was inserted'}, ignore_index=True)
    if round_new['Form of consideration'].isnull().any():
        round_new['Form of consideration'] = round_new['Form of consideration'].fillna('Cash')
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'Form of consideration',                          'Fixed Legal Team': 'Default value Cash was inserted'}, ignore_index=True)
    if round_new['Conversion'].isnull().any():
        round_new['Conversion'] = round_new['Conversion'].fillna('Fixed')
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'Conversion',                          'Fixed Legal Team': 'Default value Fixed was inserted'}, ignore_index=True)
    if round_new['No. of shares'].isnull().any():
        round_new['No. of shares'] = round_new['No. of shares'].fillna(0)
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'No. of shares',                          'Fixed Legal Team': 'Default value 0 was inserted'}, ignore_index=True)
    if round_new['Par value per share'].isnull().any():
        round_new['Par value per share'] = round_new['Par value per share'].fillna(10)
        log = log.append({'Sheet Name': 'Round Creation', 'Warning At': 'Par value per share',                          'Fixed Legal Team': 'Default value 10 was inserted'}, ignore_index=True)
    if security_new['Voting Rights'].isnull().any():
        security_new['Voting Rights'] = security_new['Voting Rights'].fillna('Yes')
        log = log.append({'Sheet Name': 'Security', 'Warning At': 'Voting Rights',                          'Fixed Legal Team': 'Default value Yes was inserted'}, ignore_index=True)
    if security_new['Dividend type'].isnull().any():
        security_new['Dividend type'] = security_new['Dividend type'].fillna('Non-Cummulative')
        log = log.append({'Sheet Name': 'Security', 'Warning At': 'Dividend type',                          'Fixed Legal Team': 'Default value Non-Cummulative was inserted'}, ignore_index=True)
    if security_new['Dividend to equity holders'].isnull().any():
        security_new['Dividend to equity holders'] = security_new['Dividend to equity holders'].fillna(
            'Priority in Payment')
        log = log.append({'Sheet Name': 'Security', 'Warning At': 'Dividend to equity holders',                          'Fixed Legal Team': 'Default value Priority in Payment was inserted'}, ignore_index=True)
    if security_new['Dividend Rate'].isnull().any():
        security_new['Dividend Rate'] = security_new['Dividend Rate'].fillna(0.001)
        log = log.append({'Sheet Name': 'Security', 'Warning At': 'Dividend Rate',                          'Fixed Legal Team': 'Default value 0.001 was inserted'}, ignore_index=True)
    if security_new['Pari Passu'].isnull().any():
        security_new['Pari Passu'] = security_new['Pari Passu'].fillna('No')
        log = log.append(
            {'Sheet Name': 'Security', 'Warning At': 'Pari Passu', 'Fixed Legal Team': 'Default value No was inserted'},            ignore_index=True)
    if security_new['Rank'].isnull().any():
        security_new['Rank'] = security_new['Rank'].fillna(1)
        log = log.append(
            {'Sheet Name': 'Security', 'Warning At': 'Rank', 'Fixed Legal Team': 'Default value 1 was inserted'},            ignore_index=True)
    if security_new['Amount'].isnull().any():
        security_new['Amount'] = security_new['Amount'].fillna(1)
        log = log.append(
            {'Sheet Name': 'Security', 'Warning At': 'Amount', 'Fixed Legal Team': 'Default value 1 was inserted'},            ignore_index=True)
    if security_new['Participation'].isnull().any():
        security_new['Participation'] = security_new['Participation'].fillna('No')
        log = log.append({'Sheet Name': 'Security', 'Warning At': 'Participation',                          'Fixed Legal Team': 'Default value No was inserted'}, ignore_index=True)
    if security_new['Cap Value'].isnull().any():
        security_new['Cap Value'] = security_new['Cap Value'].fillna(0)
        log = log.append(
            {'Sheet Name': 'Security', 'Warning At': 'Cap Value', 'Fixed Legal Team': 'Default value 0 was inserted'},            ignore_index=True)
    if security_new['Stock restructuring'].isnull().any():
        security_new['Stock restructuring'] = security_new['Stock restructuring'].fillna('Yes')
        log = log.append({'Sheet Name': 'Security', 'Warning At': 'Stock restructuring',                          'Fixed Legal Team': 'Default value Yes was inserted'}, ignore_index=True)
    if security_new['Ratio'].isnull().any():
        security_new['Ratio'] = security_new['Ratio'].fillna('1:1')
        log = log.append(
            {'Sheet Name': 'Security', 'Warning At': 'Ratio', 'Fixed Legal Team': 'Default value 1:1 was inserted'},            ignore_index=True)
    if security_new['Conversion Period'].isnull().any():
        security_new['Conversion Period'] = security_new['Conversion Period'].fillna(20)
        log = log.append({'Sheet Name': 'Security', 'Warning At': 'Conversion Period',                          'Fixed Legal Team': 'Default value 20 was inserted'}, ignore_index=True)
    if security_new['Anti- dilution'].isnull().any():
        security_new['Anti- dilution'] = security_new['Anti- dilution'].fillna('Broad based weighted average')
        log = log.append({'Sheet Name': 'Security', 'Warning At': 'Anti- dilution',                          'Fixed Legal Team': 'Default value Broad based weighted average was inserted'},                         ignore_index=True)
    a = 0
    for i in range(len(security_new)):
        if security_new['Instrument Name or Security Name'][i] == np.nan and security_new['Type of shares'][i] == 'Equity':
            security_new['Instrument Name or Security Name'][i] = 'Equity'
            a = 1
    if a == 1:
        log = log.append({'Sheet Name': 'Security', 'Warning Hissa': 'Instrument Name or Security Name','Fixed Hissa': 'Default value Common for Equity shares was inserted'}, ignore_index=True)
    for i in range(len(round_new)):
        if round_new['Type of shares'][i] == 'Equity' or round_new['Type of shares'][i] == 'Nominal' :
            round_new['Instrument Name or Security Name'][i] = 'Equity'
            a = 2
    if a == 2:
        log = log.append({'Sheet Name': 'Round Creation', 'Warning Hissa': 'Instrument Name or Security Name','Fixed Hissa': 'Default value Equity for Equity shares was inserted'}, ignore_index=True)
    round_new['Premium Per share'] = round_new['Premium Per share'].fillna(0)
    round_new['Premium Per share']=round_new['Premium Per share'].replace(",","",regex=True)
    for i in range(len(round_new)):
        try:
            round_new['Premium Per share'][i]=float(round_new['Premium Per share'][i])
        except:
            round_new['Premium Per share'][i]=0
    for i in range(len(round_new)):
        for j in range(i + 1, len(round_new)):
            if round_new['Round Name'][i] == round_new['Round Name'][j] and round_new['Name of shareholder'][i] ==  round_new['Name of shareholder'][j] and round_new['Type of shares'][i] == round_new['Type of shares'][j] and round_new['Per share price'][i]==round_new['Per share price'][j]:
                round_new['No. of shares'][i] = round_new['No. of shares'][i] + round_new['No. of shares'][j]
                round_new['Total amount invested'][i] = round_new['Total amount invested'][i] + \
                   round_new['Total amount invested'][j]
                round_new['Type of shares'][j] = np.nan
    listofroundnames=[]
    for i in range(len(round_new)):
        for j in range(i+1,len(round_new)):
            if round_new['Round Name'][i]==round_new['Round Name'][j]:
                if round_new['Per share price'][i]!=round_new['Per share price'][j]:
                    listofroundnames.append(round_new['Round Name'][j])
    listofroundnames=removeduplicate(listofroundnames)
    for i in range(len(listofroundnames)):
        log = log.append({'Sheet Name': 'Round Creation', 'Error At': 'Per share price not matching for'+str(listofroundnames[i])}, ignore_index=True)
    for i in range(len(round_new)):
        if round_new['No. of shares'][i]==0:
            round_new['Type of shares'][i] = np.nan
    round_new.dropna(subset=['Type of shares'], inplace=True)
    round_new.reset_index(inplace=True)
    if company_new['Business Details Name'].isnull().any():
        company_new['Business Details Name'] = company_new['Business Details Name'].fillna('Service')
        log = log.append({'Sheet Name': 'Company', 'Warning Hissa': 'Business Details Name',                          'Fixed Hissa': 'Default value Service was inserted'}, ignore_index=True)
    if company_new['Industry'].isnull().any():
        company_new['Industry'] = company_new['Industry'].fillna('Commercial')
        log = log.append({'Sheet Name': 'Company', 'Warning Hissa': 'Industry',                          'Fixed Hissa': 'Default value Commercial was inserted'}, ignore_index=True)
    if company_new['Nic code'].isnull().any():
        company_new['Nic code'] = company_new['Nic code'].fillna('69111')
        log = log.append(
            {'Sheet Name': 'Company', 'Warning Hissa': 'Nic code', 'Fixed Hissa': 'Default value 69111 was inserted'},            ignore_index=True)
    if founder_new['Email'].isnull().any():
        founder_new['Email'] = founder_new['Email'].fillna(company_new['Email Id'][0])
        log = log.append(
            {'Sheet Name': 'Founder', 'Warning Hissa': 'Email', 'Fixed Hissa': 'Company Email was inserted'},            ignore_index=True)
    if director_new['Email'].isnull().any():
        director_new['Email'] = director_new['Email'].fillna(company_new['Email Id'][0])
        log = log.append(
            {'Sheet Name': 'Director', 'Warning Hissa': 'Email', 'Fixed Hissa': 'Company Email was inserted'},            ignore_index=True)
    if director_new['Nationality'].isnull().any():
        director_new['Nationality'] = director_new['Nationality'].fillna('Indian')
        log = log.append({'Sheet Name': 'Director', 'Warning Hissa': 'Nationality',                          'Fixed Hissa': 'Default value Indian was inserted'}, ignore_index=True)
    before = len(secondary_new.index.tolist())
    secondary_new.dropna(subset=['Seller name'], inplace=True)
    after = len(secondary_new.index.tolist())
    if before == after:
        secondary_new.dropna(subset=['Buyer Name'], inplace=True)
    else:
        log = log.append({'Sheet Name': 'Secondary', 'Warning At': 'Secondary Sheet',                          'Fixed Legal Team': 'Seller name not found so removed row'}, ignore_index=True)
    secondary_new.reset_index(inplace=True)
    secondary_new['date of Transfer'] = pd.to_datetime(secondary_new['date of Transfer'],errors='coerce').dt.strftime('%d-%m-%Y')
    secondary_new['Shareholder Type']=secondary_new['Shareholder Type'].fillna('Others')
    round_new['Name of shareholder'].replace(r"^ +| +$", r"", regex=True, inplace=True)
    for i in range(len(secondary_new)):
        if type(secondary_new['Seller name'][i]) == str:
            secondary_new['Seller name'][i] = secondary_new['Seller name'][i].title()
    for i in range(len(secondary_new)):
        if type(secondary_new['Buyer Name'][i]) == str:
            secondary_new['Buyer Name'][i] = secondary_new['Buyer Name'][i].title()
    secondary_new['Buyer Name'].replace(r"^ +| +$", r"", regex=True, inplace=True)
    secondary_new['Seller name'].replace(r"^ +| +$", r"", regex=True, inplace=True)
    secnotthere=[]
    for i in range(len(secondary_new)):
        if secondary_new['Seller name'][i] not in round_new['Name of shareholder'].tolist():
            secnotthere.append(secondary_new['Seller name'][i])
    secnotthere=removeduplicate(secnotthere)
    finallist=[]
    if len(secnotthere)>0:
        for i in range(len(secnotthere)):
            for j in range(len(secondary_new)):
                if secnotthere[i] not in secondary_new['Buyer Name'][j]:
                    finallist.append(secnotthere[i])
    finallist=removeduplicate(finallist)
    for i in range(len(finallist)):
        log = log.append({'Sheet Name': 'Secondary', 'Error At': 'Seller name '+finallist[i]+' is not a shareholder'}, ignore_index=True)
    for i in range(len(round_new)):
        if round_new['Type of shares'][i] == 'EQUITY':
            round_new['Type of shares'][i]=round_new['Type of shares'][i].title()
    for i in range(len(secondary_new)):
        if type(secondary_new['date of Transfer'][i])==str:
            secondary_new['date of Transfer'][i] = datetime.strptime(str(secondary_new['date of Transfer'][i]),'%d-%m-%Y').date()
    founder_new['Designation']=founder_new['Designation'].fillna('Co-Founder')
    # for i in range(len(esop_new)):
    #     if type(esop_new['Financial Year'][i])==str:
    #         esop_new['Financial Year'][i] = datetime.strptime(str(esop_new['Financial Year']),'%d-%m-%Y').date()
    # for i in range(len(director_new)):
    #     if director_new['Contact Number'][i]=='-':
    #         director_new['Contact Number'][i]=np.nan
    for i in range(len(director_new)):
        if type(director_new['DIN No'][i])==str:
            director_new['DIN No'][i]=1111111
    # for i in range(len(secondary_new)):
    #     if secondary_new['Contact Number'][i]=='-':
    #         secondary_new['Contact Number'][i]=np.nan
    founder_new['Contact Number']=founder_new['Contact Number'].replace("-","",regex=True)
    round_new.insert(0, 'S.no.', range(1, 1 + len(round_new)))
    security_new.insert(0, 'S.no.', range(1, 1 + len(security_new)))
    secondary_new.insert(0, 'Sl no', range(1, 1 + len(secondary_new)))
    if len(company_new) > 0:
        for i in range(len(company_mandatory)):
            if company_new[company_mandatory[i]].isnull().any():
                log = log.append({'Sheet Name': 'Company', 'Error At': company_mandatory[i]}, ignore_index=True)
    else:
        log = log.append({'Sheet Name': 'Company', 'Error At': 'Sheet is empty'}, ignore_index=True)
    if len(founder_new) > 0:
        for i in range(len(founder_mandatory)):
            if founder_new[founder_mandatory[i]].isnull().any():
                log = log.append({'Sheet Name': 'Founder', 'Error At': founder_mandatory[i]}, ignore_index=True)
    else:
        log = log.append({'Sheet Name': 'Founder', 'Error At': 'Founder is empty'}, ignore_index=True)
    if len(director_new) > 0:
        for i in range(len(director_mandatory)):
            if director_new[director_mandatory[i]].isnull().any():
                log = log.append({'Sheet Name': 'Director', 'Error At': director_mandatory[i]}, ignore_index=True)
    else:
        log = log.append({'Sheet Name': 'Director', 'Error At': 'Director is empty'}, ignore_index=True)
    if len(round_mandatory) > 0:
        for i in range(len(round_mandatory)):
            if round_new[round_mandatory[i]].isnull().any():
                log = log.append({'Sheet Name': 'Round Creation', 'Error At': round_mandatory[i]}, ignore_index=True)
    else:
        log = log.append({'Sheet Name': 'Round Creation', 'Error At': 'Round Creation is empty'}, ignore_index=True)
    if len(secondary_new) > 0:
        for i in range(len(Sec_mandatory)):
            if (secondary_new[Sec_mandatory[i]].isnull().any()):
                log = log.append({'Sheet Name': 'Secondary', 'Error At': Sec_mandatory[i]}, ignore_index=True)
    if len(esop_new) > 0:
        for i in range(len(Esop_mandatory)):
            if (esop_new[Esop_mandatory[i]].isnull().any()):
                log = log.append({'Sheet Name': 'Esop', 'Error At': Esop_mandatory[i]}, ignore_index=True)
    roundcreationbefore2020=pd.DataFrame(columns=['S.no.', 'Name of shareholder', 'Legal Name of shareholder', 'Type of Shareholder', 'Nationality', 'ID#',         'ID type', 'Residential status', 'No. of shares', 'Type of shares', 'Round Name',         'Instrument Name or Security Name', 'Date of allotment', 'Par value per share', 'Premium Per share',         'Per share price', 'Per share price (FX) if applicable', 'Total amount invested', 'Form of consideration',         'Conversion', 'Share distinctive numbers', 'Share certificate number', 'Folio Number',         'Relevant Shareholders Agreement', 'Reference', 'Valuation as per Report(Pas 4)', 'Group Name',         'Lead Investor', 'Comments'])
    secondarybefore2020=pd.DataFrame(columns=['Sl no', 'Seller name', 'ID #', 'ID Type', 'date of Transfer', 'Type of Shares', 'Number of shares',         'Instrument Name or Security Name', 'Price per share', 'total Price', 'Share start number', 'Share end number',         'Buyer Name', 'Group Name Buyer', 'Shareholder Type', 'Registerd Address', 'Resident Status', 'Occupation',         'Nationality', 'Contact Number', 'email Id', 'Comment'])
    roundcreationbefore2020 = round_new.loc[round_new['Date of allotment']<=( pd.to_datetime('31-03-2020').date())]
    secondarybefore2020=secondary_new.loc[secondary_new['date of Transfer']<=( pd.to_datetime('31-03-2020').date())]
    if 'Secondary' not in log['Sheet Name'].tolist() and 'Date of allotment' not in log['Error At'].tolist():
        summarycaptablebefore2020=pd.DataFrame(columns=['Name','Number of equity shares','Number of preference shares','Number of debentures shares'])
        listofinvestors=[]
        for i in range(len(roundcreationbefore2020)):
            listofinvestors.append(str(roundcreationbefore2020['Name of shareholder'][i]))
        listofequityshares=[]
        listofpreferenceshares=[]
        listofdebentureshares=[]
        for i in range(len(roundcreationbefore2020)):
            if roundcreationbefore2020['Type of shares'][i]=='Equity':
                listofequityshares.append(roundcreationbefore2020['No. of shares'][i])
            else:
                listofequityshares.append(0)
        for i in range(len(roundcreationbefore2020)):
            if roundcreationbefore2020['Type of shares'][i]=='CCPS' or roundcreationbefore2020['Type of shares'][i]=='OCPS' or roundcreationbefore2020['Type of shares'][i]=='OCRPS' or roundcreationbefore2020['Type of shares'][i]=='RPS' :
                listofpreferenceshares.append(roundcreationbefore2020['No. of shares'][i])
            else:
                listofpreferenceshares.append(0) 
        for i in range(len(roundcreationbefore2020)):
            if roundcreationbefore2020['Type of shares'][i]=='CCD' or roundcreationbefore2020['Type of shares'][i]=='Note' or roundcreationbefore2020['Type of shares'][i]=='Warrant' or roundcreationbefore2020['Type of shares'][i]=='NCD' or roundcreationbefore2020['Type of shares'][i]=='Options' :
                listofdebentureshares.append(roundcreationbefore2020['No. of shares'][i])
            else:
                listofdebentureshares.append(0)
        summarycaptablebefore2020['Name']=listofinvestors
        summarycaptablebefore2020['Number of equity shares']=listofequityshares
        summarycaptablebefore2020['Number of preference shares']=listofpreferenceshares
        summarycaptablebefore2020['Number of debentures shares']=listofdebentureshares
        for i in range(len(summarycaptablebefore2020)):
            for j in range(i+1,len(summarycaptablebefore2020)):
                if summarycaptablebefore2020['Name'][i]==summarycaptablebefore2020['Name'][j]:
                    summarycaptablebefore2020['Number of equity shares'][i]=summarycaptablebefore2020['Number of equity shares'][i]+summarycaptablebefore2020['Number of equity shares'][j]
                    summarycaptablebefore2020['Number of preference shares'][i]=summarycaptablebefore2020['Number of preference shares'][i]+summarycaptablebefore2020['Number of preference shares'][j]
                    summarycaptablebefore2020['Number of debentures shares'][i]=summarycaptablebefore2020['Number of debentures shares'][i]+summarycaptablebefore2020['Number of debentures shares'][j]
                    summarycaptablebefore2020['Name'][j]=np.nan
        summarycaptablebefore2020.dropna(subset=['Name'],inplace=True)
        summarycaptablebefore2020.reset_index(inplace=True)
        listofsecondarybuyername=[]
        listofsecondarysellername=[]
        listofnumberofshares=[]
        listofsecondarytypeofshares=[]
        for i in range(len(secondarybefore2020)):
            listofsecondarybuyername.append(secondarybefore2020['Buyer Name'][i])
            listofsecondarysellername.append(secondarybefore2020['Seller name'][i])
            listofnumberofshares.append(secondarybefore2020['Number of shares'][i])
            listofsecondarytypeofshares.append(secondarybefore2020['Type of Shares'][i])
        for i in range(len(listofsecondarybuyername)):
            if listofsecondarybuyername[i] not in summarycaptablebefore2020['Name'].tolist():
                summarycaptablebefore2020.loc[len(summarycaptablebefore2020.index)] = [0,listofsecondarybuyername[i],0,0,0]
        for i in range(len(listofsecondarysellername)):
            if listofsecondarysellername[i] not in summarycaptablebefore2020['Name'].tolist():
                summarycaptablebefore2020.loc[len(summarycaptablebefore2020.index)] = [0,listofsecondarysellername[i],0,0,0]
        for i in range(len(listofsecondarybuyername)):
            for j in range(len(summarycaptablebefore2020)):
                if listofsecondarybuyername[i]== summarycaptablebefore2020['Name'][j]:
                    if listofsecondarytypeofshares[i]=='Equity':
                        summarycaptablebefore2020['Number of equity shares'][j]=summarycaptablebefore2020['Number of equity shares'][j]+listofnumberofshares[i]
                    if listofsecondarytypeofshares[i]=='CCPS' or listofsecondarytypeofshares[i]=='OCPS' or listofsecondarytypeofshares[i]=='RPS' or listofsecondarytypeofshares[i]=='OCRPS':
                        summarycaptablebefore2020['Number of preference shares'][j]=summarycaptablebefore2020['Number of preference shares'][j]+listofnumberofshares[i]
                    if listofsecondarytypeofshares[i]=='NCD' or listofsecondarytypeofshares[i]== 'Options' or listofsecondarytypeofshares[i]== 'Warrant' or listofsecondarytypeofshares[i]=='Note' or listofsecondarytypeofshares[i]=='CCD':
                        summarycaptablebefore2020['Number of debentures shares'][j]=summarycaptablebefore2020['Number of debentures shares'][j]+listofnumberofshares[i]
                if listofsecondarysellername[i]== summarycaptablebefore2020['Name'][j]:
                    if listofsecondarytypeofshares[i]=='Equity':
                        summarycaptablebefore2020['Number of equity shares'][j]=summarycaptablebefore2020['Number of equity shares'][j]-listofnumberofshares[i]
                    if listofsecondarytypeofshares[i]=='CCPS' or listofsecondarytypeofshares[i]=='OCPS' or listofsecondarytypeofshares[i]=='RPS' or listofsecondarytypeofshares[i]=='OCRPS':
                        summarycaptablebefore2020['Number of preference shares'][j]=summarycaptablebefore2020['Number of preference shares'][j]-listofnumberofshares[i]
                    if listofsecondarytypeofshares[i]=='NCD' or listofsecondarytypeofshares[i]== 'Options' or listofsecondarytypeofshares[i]== 'Warrant' or listofsecondarytypeofshares[i]=='Note' or listofsecondarytypeofshares[i]=='CCD':
                        summarycaptablebefore2020['Number of debentures shares'][j]=summarycaptablebefore2020['Number of debentures shares'][j]-listofnumberofshares[i]
        totalequityshares=0
        totalpreferenceshares=0
        totaldebentureshares=0
        for i in range(len(summarycaptablebefore2020)):
            totalequityshares+=summarycaptablebefore2020['Number of equity shares'][i]
            totalpreferenceshares+=summarycaptablebefore2020['Number of preference shares'][i]
            totaldebentureshares+=summarycaptablebefore2020['Number of debentures shares'][i]
        summarycaptablebefore2020.loc[len(summarycaptablebefore2020.index)] = [0,'Total', totalequityshares, totalpreferenceshares,totaldebentureshares]
        summarycaptablebefore2020.insert(0, 'Sl no.', range(1, 1 + len(summarycaptablebefore2020)))
        summarycaptablebefore2020=summarycaptablebefore2020[['Sl no.','Name','Number of equity shares','Number of preference shares','Number of debentures shares']]
        summarycaptable=pd.DataFrame(columns=['Name','Number of equity shares','Number of preference shares','Number of debentures shares'])
        roundcreationbefore2020=pd.DataFrame(columns=['S.no.', 'Name of shareholder', 'Legal Name of shareholder', 'Type of Shareholder', 'Nationality', 'ID#',         'ID type', 'Residential status', 'No. of shares', 'Type of shares', 'Round Name',         'Instrument Name or Security Name', 'Date of allotment', 'Par value per share', 'Premium Per share',         'Per share price', 'Per share price (FX) if applicable', 'Total amount invested', 'Form of consideration',         'Conversion', 'Share distinctive numbers', 'Share certificate number', 'Folio Number',         'Relevant Shareholders Agreement', 'Reference', 'Valuation as per Report(Pas 4)', 'Group Name',         'Lead Investor', 'Comments'])
        secondarybefore2020=pd.DataFrame(columns=['Sl no', 'Seller name', 'ID #', 'ID Type', 'date of Transfer', 'Type of Shares', 'Number of shares',         'Instrument Name or Security Name', 'Price per share', 'total Price', 'Share start number', 'Share end number',         'Buyer Name', 'Group Name Buyer', 'Shareholder Type', 'Registerd Address', 'Resident Status', 'Occupation',         'Nationality', 'Contact Number', 'email Id', 'Comment'])
        listofinvestors=[]
        for i in range(len(round_new)):
            listofinvestors.append(str(round_new['Name of shareholder'][i]))
        listofequityshares=[]
        listofpreferenceshares=[]
        listofdebentureshares=[]
        for i in range(len(round_new)):
            if round_new['Type of shares'][i]=='Equity':
                listofequityshares.append(round_new['No. of shares'][i])
            else:
                listofequityshares.append(0)
        for i in range(len(round_new)):
            if round_new['Type of shares'][i]=='CCPS' or round_new['Type of shares'][i]=='OCPS' or round_new['Type of shares'][i]=='OCRPS' or round_new['Type of shares'][i]=='RPS' :
                listofpreferenceshares.append(round_new['No. of shares'][i])
            else:
                listofpreferenceshares.append(0) 
        for i in range(len(round_new)):
            if round_new['Type of shares'][i]=='CCD' or round_new['Type of shares'][i]=='Note' or round_new['Type of shares'][i]=='Warrant' or round_new['Type of shares'][i]=='NCD' or round_new['Type of shares'][i]=='Options' :
                listofdebentureshares.append(round_new['No. of shares'][i])
            else:
                listofdebentureshares.append(0)
        summarycaptable['Name']=listofinvestors
        summarycaptable['Number of equity shares']=listofequityshares
        summarycaptable['Number of preference shares']=listofpreferenceshares
        summarycaptable['Number of debentures shares']=listofdebentureshares
        for i in range(len(summarycaptable)):
            for j in range(i+1,len(summarycaptable)):
                if summarycaptable['Name'][i]==summarycaptable['Name'][j]:
                    summarycaptable['Number of equity shares'][i]=summarycaptable['Number of equity shares'][i]+summarycaptable['Number of equity shares'][j]
                    summarycaptable['Number of preference shares'][i]=summarycaptable['Number of preference shares'][i]+summarycaptable['Number of preference shares'][j]
                    summarycaptable['Number of debentures shares'][i]=summarycaptable['Number of debentures shares'][i]+summarycaptable['Number of debentures shares'][j]
                    summarycaptable['Name'][j]=np.nan
        summarycaptable.dropna(subset=['Name'],inplace=True)
        summarycaptable.reset_index(inplace=True)
        listofsecondarybuyername=[]
        listofsecondarysellername=[]
        listofnumberofshares=[]
        listofsecondarytypeofshares=[]
        for i in range(len(secondary_new)):
            listofsecondarybuyername.append(secondary_new['Buyer Name'][i])
            listofsecondarysellername.append(secondary_new['Seller name'][i])
            listofnumberofshares.append(secondary_new['Number of shares'][i])
            listofsecondarytypeofshares.append(secondary_new['Type of Shares'][i])
        for i in range(len(listofsecondarybuyername)):
            if listofsecondarybuyername[i] not in summarycaptable['Name'].tolist():
                summarycaptable.loc[len(summarycaptable.index)] = [0,listofsecondarybuyername[i],0,0,0]
        for i in range(len(listofsecondarysellername)):
            if listofsecondarysellername[i] not in summarycaptable['Name'].tolist():
                summarycaptable.loc[len(summarycaptable.index)] = [0,listofsecondarysellername[i],0,0,0]
        for i in range(len(listofsecondarybuyername)):
            for j in range(len(summarycaptable)):
                if listofsecondarybuyername[i]== summarycaptable['Name'][j]:
                    if listofsecondarytypeofshares[i]=='Equity':
                        summarycaptable['Number of equity shares'][j]=summarycaptable['Number of equity shares'][j]+listofnumberofshares[i]
                    if listofsecondarytypeofshares[i]=='CCPS' or listofsecondarytypeofshares[i]=='OCPS' or listofsecondarytypeofshares[i]=='RPS' or listofsecondarytypeofshares[i]=='OCRPS':
                        summarycaptable['Number of preference shares'][j]=summarycaptable['Number of preference shares'][j]+listofnumberofshares[i]
                    if listofsecondarytypeofshares[i]=='NCD' or listofsecondarytypeofshares[i]== 'Options' or listofsecondarytypeofshares[i]== 'Warrant' or listofsecondarytypeofshares[i]=='Note' or listofsecondarytypeofshares[i]=='CCD':
                        summarycaptable['Number of debentures shares'][j]=summarycaptable['Number of debentures shares'][j]+listofnumberofshares[i]
                if listofsecondarysellername[i]== summarycaptable['Name'][j]:
                    if listofsecondarytypeofshares[i]=='Equity':
                        summarycaptable['Number of equity shares'][j]=summarycaptable['Number of equity shares'][j]-listofnumberofshares[i]
                    if listofsecondarytypeofshares[i]=='CCPS' or listofsecondarytypeofshares[i]=='OCPS' or listofsecondarytypeofshares[i]=='RPS' or listofsecondarytypeofshares[i]=='OCRPS':
                        summarycaptable['Number of preference shares'][j]=summarycaptable['Number of preference shares'][j]-listofnumberofshares[i]
                    if listofsecondarytypeofshares[i]=='NCD' or listofsecondarytypeofshares[i]== 'Options' or listofsecondarytypeofshares[i]== 'Warrant' or listofsecondarytypeofshares[i]=='Note' or listofsecondarytypeofshares[i]=='CCD':
                        summarycaptable['Number of debentures shares'][j]=summarycaptable['Number of debentures shares'][j]-listofnumberofshares[i]
        totalequityshares=0
        totalpreferenceshares=0
        totaldebentureshares=0
        for i in range(len(summarycaptable)):
            totalequityshares+=summarycaptable['Number of equity shares'][i]
            totalpreferenceshares+=summarycaptable['Number of preference shares'][i]
            totaldebentureshares+=summarycaptable['Number of debentures shares'][i]
        summarycaptable.loc[len(summarycaptable.index)] = [0,'Total', totalequityshares, totalpreferenceshares,totaldebentureshares]
        summarycaptable.insert(0, 'Sl no.', range(1, 1 + len(summarycaptable)))
        summarycaptable=summarycaptable[['Sl no.','Name','Number of equity shares','Number of preference shares','Number of debentures shares']]
    for i in range(len(security_new)):
        if security_new['Type of Shares'][i] == 'Equity':
            security_new['Voting Rights'][i] = np.nan
            security_new['Dividend type'][i] = np.nan
            security_new['Dividend to equity holders'][i] = np.nan
            security_new['Dividend Rate'][i] = np.nan
            security_new['Pari Passu'][i] = np.nan
            security_new['Rank'][i] = np.nan
            security_new['Amount'][i] = np.nan
            security_new['Participation'][i] = np.nan
            security_new['Cap Value'][i] = np.nan
            security_new['Ratio'][i] = np.nan
            security_new['Conversion Period'][i] = np.nan
            security_new['Anti- dilution'][i] = np.nan
            security_new['Stock restructuring'][i] = np.nan
    roundname = []
    conversionofshares = 0
    for i in range(len(round_new['Round Name'])):
        roundname.append(str(round_new['Round Name'][i]).lower())
        if roundname[i].find("conversion") != -1:
            conversionofshares = 1
    if company_new['Par Value'].isnull().any():
        dictionary.update({'company Par value missing': 'Yes'})
    else:
        dictionary.update({'company Par value missing': 'No'})
    if round_new['Round Name'].isnull().any():
        dictionary.update({'Round Creation Round Name missing': 'Yes'})
    else:
        dictionary.update({'Round Creation Round Name missing': 'No'})
    if round_new['Per share price'].isnull().any():
        dictionary.update({'Round Creation PPS missing': 'Yes'})
    else:
        dictionary.update({'Round Creation PPS missing': 'No'})
    if conversionofshares == 1:
        dictionary.update({'Conversion of shares is there?': 'Yes'})
    else:
        dictionary.update({'Conversion of shares is there?': 'No'})
    # for i in range(len(company_new)):
    #     if type(company_new['Paid up Capital (INR)'][i])== str: 
    #         company_new['Paid up Capital (INR)'][i]=company_new['Paid up Capital (INR)'][i].replace(',','',regex=True)
    paidup_match = False
    company_new=company_new.replace(',','',regex=True)
    company_new['Paid up Capital (INR)']=company_new['Paid up Capital (INR)'].astype(int)
    if sum == company_new['Paid up Capital (INR)'][0]:
        dictionary.update({'Paidup Matched': 'Yes'})
        paidup_match = True
    else:
        dictionary.update({'Paidup Matched': 'No'})
        paidup_match = False
    if len(round_new) > 50:
        dictionary.update({'Round Creation Less than 50 entries': 'No'})
    else:
        dictionary.update({'Round Creation Less than 50 entries': 'Yes'})
    partlypaid=False
    partypaidcheckvalue=company_new['Paid up Capital (INR)'][0] % 10
    if partypaidcheckvalue!=0:
        dictionary.update({'Partly Paid':'Yes'})
        partlypaid=True
    for i in range(len(round_new)):
        for j in range(len(round_new)):
            if round_new['Type of shares'][i]=='Equity' or 'EQUITY':
                if round_new['Type of shares'][j]==round_new['Type of shares'][i]:
                    if round_new['Par value per share'][i]!=round_new['Par value per share'][j]:
                        dictionary.update({'Split is there?':'Yes'})            
    if 'Secondary' in log['Sheet Name'].tolist():
        dictionary.update({'Secondary Errors':'Yes'})
    numberoferrors=log['Error At'].count()
    dictionary.update({'Numberoferrors':numberoferrors})
    status=''
    if log['Error At'].count() == 0 and log['Sheet Name'].count() == 0:
        dictionary.update({'Status': 'Clean'})
    elif (log['Fixed Legal Team'].count() + log['Fixed Hissa'].count()) == log['Sheet Name'].count() and log[
        'Error At'].count() == 0 and log['Sheet Name'].count() > 0 and paidup_match == True:
        dictionary.update({'Status': 'Fixed'})
        status='fixed'
    elif (log['Fixed Legal Team'].count() + log['Fixed Hissa'].count()) == log['Sheet Name'].count() and log[
        'Error At'].count() == 0 and log['Sheet Name'].count() > 0 and paidup_match == False:
        dictionary.update({'Status': 'Paidup'})
        status='paidup'
    else:
        dictionary.update({'Status': 'Error'})
        status='error'
    company_new = company_new[
        ['CIN', 'Company Name', 'Company Category', 'Class of Company', 'Authorised Capital (INR) as of date',         'Authorised Capital at Incorporation', 'Paid up Capital (INR)', 'Date of Incorporation', 'Registered Address',         'Address other than R/o where all or any books of account and papers are maintained', 'Email Id',         'Whether Listed or not', 'Company Status(for efiling)', 'Dipp certificate number', 'Par Value',         'Business Details Name', 'Industry', 'Nic code', 'Company Website']]
    founder_new = founder_new[['Name', 'Email', 'Nationality', 'Designation', 'Contact Number', 'Address', 'Dob']]
    director_new = director_new[
        ['DIN No', 'Name', 'Date of Appointment', 'Date of board resolution for appointment', 'Residential status',         'Type of Director', 'Date of resignation', 'Email', 'Nationality', 'dateOfBirth', 'Address']]
    round_new = round_new[
        ['S.no.', 'Name of shareholder', 'Legal Name of shareholder', 'Type of Shareholder', 'Nationality', 'ID#',         'ID type', 'Residential status', 'No. of shares', 'Type of shares', 'Round Name',         'Instrument Name or Security Name', 'Date of allotment', 'Par value per share', 'Premium Per share',         'Per share price', 'Per share price (FX) if applicable', 'Total amount invested', 'Form of consideration',         'Conversion', 'Share distinctive numbers', 'Share certificate number', 'Folio Number',         'Relevant Shareholders Agreement', 'Reference', 'Valuation as per Report(Pas 4)', 'Group Name',         'Lead Investor', 'Comments']]
    security_new = security_new[
        ['S.no.', 'Round Name', 'Price per share', 'No. of Shares in this round', 'Total Premoney Shares',         'Pre Money Valuation', 'Post Money Valuation', 'Instrument Name or Security Name', 'Type of Shares',         'Par value', 'Voting Rights', 'Dividend type', 'Dividend to equity holders', 'Dividend Rate', 'Pari Passu',         'Rank', 'Amount', 'Participation', 'Cap Value', 'Ratio', 'Conversion Period',         'Valuation adjustment based on convertible', 'Conversion Forumula', 'Discount/Conversion terms',         'Valuation Cap (INR)', 'Anti- dilution', 'Stock restructuring', 'Number of warrants', 'Exercise price',         'Upfront amount', 'Lock-in period', 'Exercise Period', 'Redemption Amount', 'Note Period', 'Note Amount',         'Note Interest', 'Note Payment', 'Comment']]
    secondary_new = secondary_new[
        ['Sl no', 'Seller name', 'ID #', 'ID Type', 'date of Transfer', 'Type of Shares', 'Number of shares',         'Instrument Name or Security Name', 'Price per share', 'total Price', 'Share start number', 'Share end number',         'Buyer Name', 'Group Name Buyer', 'Shareholder Type', 'Registerd Address', 'Resident Status', 'Occupation',         'Nationality', 'Contact Number', 'email Id', 'Comment']]
    esop_new = esop_new[
        ['Financial Year', 'Pool', 'Granted', 'Vested', 'Exercised', 'Forfeited', 'Available for Grant']]
    if partlypaid==True:
        error='Error_folder'
        answer=os.path.normpath(basepath + os.sep + os.pardir)
        error_path = os.path.join(answer, error)
        isdir_error = os.path.isdir(error_path)
        if isdir_error==True:
            a=0
        else:
            os.mkdir(error_path)
        error_path = error_path + '\\'
        if partlypaid==True:
            PartlyPaid='PartlyPaid'
            partlypaid_path = os.path.join(error_path, PartlyPaid)
            isdir_partlypaid = os.path.isdir(partlypaid_path)
            if isdir_partlypaid==True:
                a=0
            else:
                os.mkdir(partlypaid_path)
            partlypaid_path = partlypaid_path + '\\'
        with pd.ExcelWriter(path=partlypaid_path+bFile, engine='xlsxwriter', date_format='dd-mm-yyyy',                            engine_kwargs={'options': {'strings_to_numbers': False}}) as writer:
            log.to_excel(writer, sheet_name='LOG', index=False)
            for column in log:
                column_width = max(log[column].astype(str).map(len).max(), len(column))
                col_idx = log.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['LOG'].set_column(col_idx, col_idx, column_width)
            if 'Secondary' not in log['Sheet Name'].tolist() and 'Date of allotment' not in log['Error At'].tolist():
                summarycaptable.to_excel(writer, sheet_name='summarycaptable', index=False)
                for column in summarycaptable:
                    column_width = max(summarycaptable[column].astype(str).map(len).max(), len(column))
                    col_idx = summarycaptable.columns.get_loc(column)
                    if column_width in range(1, 1000):
                        writer.sheets['summarycaptable'].set_column(col_idx, col_idx, column_width)
                summarycaptablebefore2020.to_excel(writer, sheet_name='summarycaptablebefore2020', index=False)
                for column in summarycaptablebefore2020:
                    column_width = max(summarycaptablebefore2020[column].astype(str).map(len).max(), len(column))
                    col_idx = summarycaptablebefore2020.columns.get_loc(column)
                    if column_width in range(1, 1000):
                        writer.sheets['summarycaptablebefore2020'].set_column(col_idx, col_idx, column_width)
            User.to_excel(writer, sheet_name='User', index=False)
            for column in User:
                column_width = max(User[column].astype(str).map(len).max(), len(column))
                col_idx = User.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['User'].set_column(col_idx, col_idx, column_width)
            company_new.to_excel(writer, sheet_name='Company', index=False)
            for column in company_new:
                column_width = max(company_new[column].astype(str).map(len).max(), len(column))
                col_idx = company_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Company'].set_column(col_idx, col_idx, column_width)
            founder_new.to_excel(writer, sheet_name='Founders', index=False)
            for column in founder_new:
                column_width = max(founder_new[column].astype(str).map(len).max(), len(column))
                col_idx = founder_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Founders'].set_column(col_idx, col_idx, column_width)
            director_new.to_excel(writer, sheet_name='Director', index=False)
            for column in director_new:
                column_width = max(director_new[column].astype(str).map(len).max(), len(column))
                col_idx = director_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Director'].set_column(col_idx, col_idx, column_width)
            round_new.to_excel(writer, sheet_name='Round Creation', index=False)
            for column in round_new:
                column_width = max(round_new[column].astype(str).map(len).max(), len(column))
                col_idx = round_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Round Creation'].set_column(col_idx, col_idx, column_width)
            security_new.to_excel(writer, sheet_name='Security', index=False)
            for column in security_new:
                column_width = max(security_new[column].astype(str).map(len).max(), len(column))
                col_idx = security_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Security'].set_column(col_idx, col_idx, column_width)
            secondary_new.to_excel(writer, sheet_name='Secondary', index=False)
            for column in secondary_new:
                column_width = max(secondary_new[column].astype(str).map(len).max(), len(column))
                col_idx = secondary_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Secondary'].set_column(col_idx, col_idx, column_width)
            esop_new.to_excel(writer, sheet_name='ESOP', index=False)
            for column in esop_new:
                column_width = max(esop_new[column].astype(str).map(len).max(), len(column))
                col_idx = esop_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['ESOP'].set_column(col_idx, col_idx, column_width)
    if status=='fixed':
        fixed='Fixed'
        answer=os.path.normpath(basepath + os.sep + os.pardir)
        fixed_path = os.path.join(answer, fixed)
        isdir_fixed = os.path.isdir(fixed_path)
        if isdir_fixed==True:
            a=0
        else:
            os.mkdir(fixed_path)
        fixed_path = fixed_path + '\\'
        with pd.ExcelWriter(path=fixed_path+bFile, engine='xlsxwriter', date_format='dd-mm-yyyy',                            engine_kwargs={'options': {'strings_to_numbers': False}}) as writer:
            log.to_excel(writer, sheet_name='LOG', index=False)
            for column in log:
                column_width = max(log[column].astype(str).map(len).max(), len(column))
                col_idx = log.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['LOG'].set_column(col_idx, col_idx, column_width)
            if 'Secondary' not in log['Sheet Name'].tolist() and 'Date of allotment' not in log['Error At'].tolist(): 
                summarycaptable.to_excel(writer, sheet_name='summarycaptable', index=False)
                for column in summarycaptable:
                    column_width = max(summarycaptable[column].astype(str).map(len).max(), len(column))
                    col_idx = summarycaptable.columns.get_loc(column)
                    if column_width in range(1, 1000):
                        writer.sheets['summarycaptable'].set_column(col_idx, col_idx, column_width)
                summarycaptablebefore2020.to_excel(writer, sheet_name='summarycaptablebefore2020', index=False)
                for column in summarycaptablebefore2020:
                    column_width = max(summarycaptablebefore2020[column].astype(str).map(len).max(), len(column))
                    col_idx = summarycaptablebefore2020.columns.get_loc(column)
                    if column_width in range(1, 1000):
                        writer.sheets['summarycaptablebefore2020'].set_column(col_idx, col_idx, column_width)
            User.to_excel(writer, sheet_name='User', index=False)
            for column in User:
                column_width = max(User[column].astype(str).map(len).max(), len(column))
                col_idx = User.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['User'].set_column(col_idx, col_idx, column_width)
            company_new.to_excel(writer, sheet_name='Company', index=False)
            for column in company_new:
                column_width = max(company_new[column].astype(str).map(len).max(), len(column))
                col_idx = company_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Company'].set_column(col_idx, col_idx, column_width)
            founder_new.to_excel(writer, sheet_name='Founders', index=False)
            for column in founder_new:
                column_width = max(founder_new[column].astype(str).map(len).max(), len(column))
                col_idx = founder_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Founders'].set_column(col_idx, col_idx, column_width)
            director_new.to_excel(writer, sheet_name='Director', index=False)
            for column in director_new:
                column_width = max(director_new[column].astype(str).map(len).max(), len(column))
                col_idx = director_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Director'].set_column(col_idx, col_idx, column_width)
            round_new.to_excel(writer, sheet_name='Round Creation', index=False)
            for column in round_new:
                column_width = max(round_new[column].astype(str).map(len).max(), len(column))
                col_idx = round_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Round Creation'].set_column(col_idx, col_idx, column_width)
            security_new.to_excel(writer, sheet_name='Security', index=False)
            for column in security_new:
                column_width = max(security_new[column].astype(str).map(len).max(), len(column))
                col_idx = security_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Security'].set_column(col_idx, col_idx, column_width)
            secondary_new.to_excel(writer, sheet_name='Secondary', index=False)
            for column in secondary_new:
                column_width = max(secondary_new[column].astype(str).map(len).max(), len(column))
                col_idx = secondary_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['Secondary'].set_column(col_idx, col_idx, column_width)
            esop_new.to_excel(writer, sheet_name='ESOP', index=False)
            for column in esop_new:
                column_width = max(esop_new[column].astype(str).map(len).max(), len(column))
                col_idx = esop_new.columns.get_loc(column)
                if column_width in range(1, 1000):
                    writer.sheets['ESOP'].set_column(col_idx, col_idx, column_width)
        prob = os.path.join(answer, "Fixed_original")
        isdir_prob = os.path.isdir(prob)
        if isdir_prob == True:
            x = 0
        else:
            os.mkdir(prob)
        original = os.path.join(basepath, bFile)
        target = os.path.join(prob, bFile)
        shutil.move(original, target)
    if status=='error' or status == 'paidup':
        error='Error_folder'
        answer=os.path.normpath(basepath + os.sep + os.pardir)
        error_path = os.path.join(answer, error)
        isdir_error = os.path.isdir(error_path)
        if isdir_error==True:
            a=0
        else:
            os.mkdir(error_path)
        error_path = error_path + '\\'
        if (status=='error' or status=='paidup') and paidup_match==False:
            paidup="Paidup"
            dictionary.update({'Status': 'Paidup'})
            paidup_path = os.path.join(error_path, paidup)
            isdir_paidup = os.path.isdir(paidup_path)
            if isdir_paidup==True:
                a=0
            else:
                os.mkdir(paidup_path)
            paidup_path = paidup_path + '\\'
            if paidup_match==False:
                with pd.ExcelWriter(path=paidup_path + bFile, engine='xlsxwriter', date_format='dd-mm-yyyy',                        engine_kwargs={'options': {'strings_to_numbers': False}}) as writer:
                    log.to_excel(writer, sheet_name='LOG', index=False)
                    for column in log:
                        column_width = max(log[column].astype(str).map(len).max(), len(column))
                        col_idx = log.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['LOG'].set_column(col_idx, col_idx, column_width)
                    if 'Secondary' not in log['Sheet Name'].tolist() and 'Date of allotment' not in log['Error At'].tolist(): 
                        summarycaptable.to_excel(writer, sheet_name='summarycaptable', index=False)
                        for column in summarycaptable:
                            column_width = max(summarycaptable[column].astype(str).map(len).max(), len(column))
                            col_idx = summarycaptable.columns.get_loc(column)
                            if column_width in range(1, 1000):
                                writer.sheets['summarycaptable'].set_column(col_idx, col_idx, column_width)
                        summarycaptablebefore2020.to_excel(writer, sheet_name='summarycaptablebefore2020', index=False)
                        for column in summarycaptablebefore2020:
                            column_width = max(summarycaptablebefore2020[column].astype(str).map(len).max(), len(column))
                            col_idx = summarycaptablebefore2020.columns.get_loc(column)
                            if column_width in range(1, 1000):
                                writer.sheets['summarycaptablebefore2020'].set_column(col_idx, col_idx, column_width)
                    User.to_excel(writer, sheet_name='User', index=False)
                    for column in User:
                        column_width = max(User[column].astype(str).map(len).max(), len(column))
                        col_idx = User.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['User'].set_column(col_idx, col_idx, column_width)
                    company_new.to_excel(writer, sheet_name='Company', index=False)
                    for column in company_new:
                        column_width = max(company_new[column].astype(str).map(len).max(), len(column))
                        col_idx = company_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Company'].set_column(col_idx, col_idx, column_width)
                    founder_new.to_excel(writer, sheet_name='Founders', index=False)
                    for column in founder_new:
                        column_width = max(founder_new[column].astype(str).map(len).max(), len(column))
                        col_idx = founder_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Founders'].set_column(col_idx, col_idx, column_width)
                    director_new.to_excel(writer, sheet_name='Director', index=False)
                    for column in director_new:
                        column_width = max(director_new[column].astype(str).map(len).max(), len(column))
                        col_idx = director_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Director'].set_column(col_idx, col_idx, column_width)
                    round_new.to_excel(writer, sheet_name='Round Creation', index=False)
                    for column in round_new:
                        column_width = max(round_new[column].astype(str).map(len).max(), len(column))
                        col_idx = round_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Round Creation'].set_column(col_idx, col_idx, column_width)
                    security_new.to_excel(writer, sheet_name='Security', index=False)
                    for column in security_new:
                        column_width = max(security_new[column].astype(str).map(len).max(), len(column))
                        col_idx = security_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Security'].set_column(col_idx, col_idx, column_width)
                    secondary_new.to_excel(writer, sheet_name='Secondary', index=False)
                    for column in secondary_new:
                        column_width = max(secondary_new[column].astype(str).map(len).max(), len(column))
                        col_idx = secondary_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Secondary'].set_column(col_idx, col_idx, column_width)
                    esop_new.to_excel(writer, sheet_name='ESOP', index=False)
                    for column in esop_new:
                        column_width = max(esop_new[column].astype(str).map(len).max(), len(column))
                        col_idx = esop_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['ESOP'].set_column(col_idx, col_idx, column_width)
        if (status=='error' or status=='paidup') and paidup_match==True:
            error="Error"
            dictionary.update({'Status': 'Error'})
            finalerror_path = os.path.join(error_path, error)
            isdir_finalerror = os.path.isdir(finalerror_path)
            if isdir_finalerror==True:
                a=0
            else:
                os.mkdir(finalerror_path)
            finalerror_path = finalerror_path + '\\'
            with pd.ExcelWriter(path=finalerror_path + bFile, engine='xlsxwriter', date_format='dd-mm-yyyy',                        engine_kwargs={'options': {'strings_to_numbers': False}}) as writer:
                    log.to_excel(writer, sheet_name='LOG', index=False)
                    for column in log:
                        column_width = max(log[column].astype(str).map(len).max(), len(column))
                        col_idx = log.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['LOG'].set_column(col_idx, col_idx, column_width)
                    
                    if 'Secondary' not in log['Sheet Name'].tolist() and 'Date of allotment' not in log['Error At'].tolist(): 
                        summarycaptable.to_excel(writer, sheet_name='summarycaptable', index=False)
                        for column in summarycaptable:
                            column_width = max(summarycaptable[column].astype(str).map(len).max(), len(column))
                            col_idx = summarycaptable.columns.get_loc(column)
                            if column_width in range(1, 1000):
                                writer.sheets['summarycaptable'].set_column(col_idx, col_idx, column_width)
                        summarycaptablebefore2020.to_excel(writer, sheet_name='summarycaptablebefore2020', index=False)
                        for column in summarycaptablebefore2020:
                            column_width = max(summarycaptablebefore2020[column].astype(str).map(len).max(), len(column))
                            col_idx = summarycaptablebefore2020.columns.get_loc(column)
                            if column_width in range(1, 1000):
                                writer.sheets['summarycaptablebefore2020'].set_column(col_idx, col_idx, column_width)
                    User.to_excel(writer, sheet_name='User', index=False)
                    for column in User:
                        column_width = max(User[column].astype(str).map(len).max(), len(column))
                        col_idx = User.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['User'].set_column(col_idx, col_idx, column_width)
                    company_new.to_excel(writer, sheet_name='Company', index=False)
                    for column in company_new:
                        column_width = max(company_new[column].astype(str).map(len).max(), len(column))
                        col_idx = company_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Company'].set_column(col_idx, col_idx, column_width)
                    founder_new.to_excel(writer, sheet_name='Founders', index=False)
                    for column in founder_new:
                        column_width = max(founder_new[column].astype(str).map(len).max(), len(column))
                        col_idx = founder_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Founders'].set_column(col_idx, col_idx, column_width)
                    director_new.to_excel(writer, sheet_name='Director', index=False)
                    for column in director_new:
                        column_width = max(director_new[column].astype(str).map(len).max(), len(column))
                        col_idx = director_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Director'].set_column(col_idx, col_idx, column_width)
                    round_new.to_excel(writer, sheet_name='Round Creation', index=False)
                    for column in round_new:
                        column_width = max(round_new[column].astype(str).map(len).max(), len(column))
                        col_idx = round_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Round Creation'].set_column(col_idx, col_idx, column_width)
                    security_new.to_excel(writer, sheet_name='Security', index=False)
                    for column in security_new:
                        column_width = max(security_new[column].astype(str).map(len).max(), len(column))
                        col_idx = security_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Security'].set_column(col_idx, col_idx, column_width)
                    secondary_new.to_excel(writer, sheet_name='Secondary', index=False)
                    for column in secondary_new:
                        column_width = max(secondary_new[column].astype(str).map(len).max(), len(column))
                        col_idx = secondary_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['Secondary'].set_column(col_idx, col_idx, column_width)
                    esop_new.to_excel(writer, sheet_name='ESOP', index=False)
                    for column in esop_new:
                        column_width = max(esop_new[column].astype(str).map(len).max(), len(column))
                        col_idx = esop_new.columns.get_loc(column)
                        if column_width in range(1, 1000):
                            writer.sheets['ESOP'].set_column(col_idx, col_idx, column_width)
    print("{}{}{}".format("End converting Process for <", bFile, ">", fileType))
    print("--------------------------------------------------------------------")
    return basepath, dictionary
def getFileType(inFilePath):
    filePath, extn = os.path.splitext(inFilePath)
    fileType = 'UNKNOWN'
    # print(filePath)
    # print(extn)
    # Supported file formats - Stored in a DICTONARY
    typesDic = {
        '.XLS': 'XLS',        '.XLSX': 'XLSX'
    }
    for mytypeDic in typesDic:
        # print(mytypeDic)
        # print(extn)
        if mytypeDic.casefold() == extn.casefold():
            fileType = typesDic[mytypeDic]
            break  # Match found
    return fileType
def parseCmdLine(argv):
    infile = ''
    parser = argparse.ArgumentParser(description='HISSA Onboarding conversion demon ...')
    parser.add_argument('indir', type=str, help='Input dir for older formatted Hissa XLS files')
    args = parser.parse_args()
    infile = args.indir
    return infile
# -----------------------------------------------------------------------
# Function   :  Main
# Description:  This is where the program starts
#
# ----------------------------------------------------------------------
if __name__ == "__main__":
    fileType = ''
    appendinglist = []
    scanDir = parseCmdLine(sys.argv[1:])
    if (os.path.isdir(scanDir) == False):
        print("{}{}{}".format("ERROR: Input Dir Path ", scanDir, " DOES NOT EXIST !!!"))
    # Read files in the directory !!!!
    basepath = scanDir
    for bFile in os.listdir(basepath):
        if os.path.isfile(os.path.join(basepath, bFile)):
            fullPath = os.path.join(basepath, bFile)
            fileType = getFileType(bFile)
            if fileType == 'UNKNOWN':
                print('ERROR: UNRECOGNISED EXTENSION !!!!!!!!!!!')
            # print("------------------------------------------------------------")
            # print("{}{}{}".format("Start converting Process for - ",bFile, "-", fileType))
            # print("------------------------------------------------------------")
            # Call the conversion API  .
            
            
            try:
                output, dictionary = start_conversionProcess(fullPath, bFile, fileType)
                appendinglist.append(dictionary)
            except:
                prob= os.path.join(basepath, "Problem")
                isdir_prob = os.path.isdir(prob)
                if isdir_prob == True:
                    x = 0
                else:
                    os.mkdir(prob)
                original=os.path.join(basepath,bFile)
                target=os.path.join(prob,bFile)
                shutil.move(original,target)
    dashboard = pd.DataFrame(columns=range(14))
    columnnames = list(appendinglist[0].keys())
    dashboard.columns = columnnames
    for i in range(len(appendinglist)):
        dashboard.loc[len(dashboard.index)] = list(appendinglist[i].values())
    dashboard.insert(0, 'Sl no.', range(1, 1 + len(dashboard)))
    now = datetime.now()
    dt_string = now.strftime("%d_%m_%Y_%H_%M_%S")
    path=os.path.normpath(basepath + os.sep + os.pardir)
    writer = pd.ExcelWriter(path + "\Dashboard_" + str(dt_string) + ".xlsx")
    dashboard.to_excel(writer, sheet_name="Dashboard", index=False)
    worksheet = writer.sheets['Dashboard']
    for i in range(len(dashboard)):
        worksheet.data_validation('O'+str(i+2), {'validate': 'list','source': ['Assigned',' Received']})
    for column in dashboard:
        column_length = max(dashboard[column].astype(str).map(len).max(), len(column))
        col_idx = dashboard.columns.get_loc(column)
        writer.sheets["Dashboard"].set_column(col_idx, col_idx, column_length)
    writer.save()
    print("".format())